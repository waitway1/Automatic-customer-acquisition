#!/usr/bin/env python3
from __future__ import annotations

import email
import imaplib
import json
import base64
import hashlib
import os
import re
import shutil
import subprocess
import sys
import threading
import time
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from email.header import decode_header
from email.utils import parsedate_to_datetime
from html import escape, unescape
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from socketserver import ThreadingMixIn
from typing import Any
from urllib.parse import parse_qs, quote_plus, unquote, urlparse
from urllib.request import Request, urlopen

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
STATIC_DIR = ROOT / "static"
RUNTIME_DIR = ROOT / "runtime"
CONFIG_PATH = ROOT / "config.json"
LOCAL_CONFIG_PATH = ROOT / "config.local.json"
LOG_PATH = RUNTIME_DIR / "events.log"
TASKS_PATH = RUNTIME_DIR / "tasks.json"
INTERVENTION_PATH = RUNTIME_DIR / "intervention.json"
BOUNCE_STATE_PATH = RUNTIME_DIR / "bounce_state.json"
MAIL_STATUS_PATH = RUNTIME_DIR / "mail_status.json"
DAILY_INVALID_PATH = RUNTIME_DIR / "daily_invalid.json"
MAIL_MESSAGES_PATH = RUNTIME_DIR / "mail_messages.json"
GENERATED_CONFIG_DIR = ROOT / "work" / "campaign_configs"
ASSET_DIR = ROOT / "work" / "assets"

EMAIL_RE = re.compile(r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}", re.I)
URL_RE = re.compile(r"https?://[^\s<>)\"']+", re.I)
ACCESSORY_TERMS = (
    "accessory",
    "accessories",
    "parts",
    "shop",
    "store",
    "4x4",
    "offroad",
    "off-road",
    "aftermarket",
    "upgrade",
    "body kit",
    "side step",
    "roof rack",
    "fender",
    "canopy",
    "tonneau",
    "bumper",
    "bull bar",
    "grille",
    "floor mat",
    "floor mats",
    "floor-mat",
    "floor-mats",
    "cross bar",
    "cross bars",
    "cross-bar",
    "cross-bars",
    "mud flap",
    "mud flaps",
    "mud-flap",
    "mud-flaps",
    "trailer hitch",
    "trailer-hitch",
    "spare tire",
    "spare-tire",
    "spare wheel",
    "spare-wheel",
)
SHOP_INTENT_TERMS = (
    "shop",
    "store",
    "dealer",
    "distributor",
    "supplier",
    "wholesale",
    "manufacturer",
    "factory",
    "contact",
    "about us",
    "online",
    "cart",
    "add to cart",
    "buy",
    "product",
    "products",
    "collection",
    "collections",
    "catalog",
    "catalogue",
    "quote",
    "quotation",
    "shipping",
    "delivery",
    "accessories",
    "parts",
    "4x4",
    "offroad",
    "off-road",
)
LEAD_SEARCH_SUFFIXES = (
    "accessories shop email",
    "accessories dealer contact",
    "parts store email",
    "4x4 accessories distributor email",
    "body kit supplier contact",
    "side steps supplier email",
    "roof rack dealer email",
    "offroad accessories wholesale email",
    "aftermarket parts contact",
    "online shop contact",
    "accessories online shop",
    "accessories wholesale contact",
    "parts distributor contact",
    "email contact",
)
INVALID_SHEET_TOKENS = (
    "邮箱失效",
    "失效邮箱",
    "邮件失效",
    "无效邮箱",
    "失效",
    "退信",
    "退回",
    "投递失败",
    "invalid",
    "bounce",
    "bounced",
    "undeliver",
)
FOLLOWUP_SHEET_TOKENS = (
    "跟进",
    "回复",
    "已回复",
    "客户回复",
    "有意向",
    "意向客户",
    "待人工",
    "待跟进",
    "follow",
    "followup",
    "follow-up",
    "reply",
    "replies",
    "interested",
)
FOLLOWUP_REPLY_TOKENS = (
    "回复",
    "已回复",
    "客户回复",
    "跟进回复",
    "有意向",
    "意向",
    "reply",
    "replies",
    "interested",
)
BAD_EMAIL_PREFIXES = (
    "noreply@",
    "no-reply@",
    "donotreply@",
    "mailer-daemon@",
    "webmaster@",
    "postmaster@",
    "hostmaster@",
)
BAD_HOSTS = {
    "youtube.com",
    "facebook.com",
    "instagram.com",
    "amazon.com",
    "alibaba.com",
    "aliexpress.com",
    "ebay.com",
    "reddit.com",
    "walmart.com",
    "etsy.com",
    "google.com",
    "maps.google.com",
    "bing.com",
    "duckduckgo.com",
    "hyundai.com",
    "hyundaiusa.com",
    "hyundaicanada.com",
    "hyundai.news",
    "jetour.com",
    "jetourglobal.com",
    "jetourauto.com",
    "cheryinternational.com",
    "byd.com",
    "made-in-china.com",
    "cautop.com",
    "shop.app",
    "gov.uk",
    "caranddriver.com",
    "techradar.com",
    "baidu.com",
    "baike.baidu.com",
    "news.cn",
    "stats.gov.cn",
    "hyundai.com.cn",
    "jetour.com.cn",
    "jetour.ltd",
    "myparts.car",
    "oempartsonline.com",
    "w3.org",
    "schemas.live.com",
    "autohome.com.cn",
    "jetour.com.hk",
    "jetourautophilippines.com",
    "jetourbr.com",
    "jetour.co.id",
    "jetoursoueast.mx",
}
BAD_HOST_SUFFIXES = (
    "google.com",
    "googleusercontent.com",
    "youtube.com",
    "facebook.com",
    "instagram.com",
    "amazon.com",
    "ebay.com",
    "alibaba.com",
    "aliexpress.com",
    "baidu.com",
    "news.cn",
    "stats.gov.cn",
    "hyundai.com",
    "jetour.com",
    "jetourglobal.com",
    "jetourauto.com",
    "cheryinternational.com",
    "w3.org",
    "schemas.live.com",
    "autohome.com.cn",
    "powerthesaurus.org",
    "merriam-webster.com",
    "wordhippo.com",
    "thesaurus.com",
    "thesaurus.plus",
    "antonym.com",
    "wordhelp.com",
    "grammardesk.com",
    "zhihu.com",
)
BAD_URL_FRAGMENTS = (
    "srsltid=",
    ".css",
    ".js",
    ".ico",
    ".png",
    ".jpg",
    ".jpeg",
    ".gif",
    ".svg",
    "/search?",
    "/maps/",
    "/news/",
    "/blog/",
    "/article/",
    "/review/",
    "/reviews/",
    "/forum/",
    "/forums/",
    "/video/",
    "/videos/",
    "/privacy",
    "/terms",
    "/terms-of-use",
    "/terms-and-conditions",
    "/accessibility",
    "/careers",
    "/jobs",
    "/service",
    "/services",
    "honda-of-santa-fe",
    "santa-fe-online-store",
    "thecollectivesantafe",
    "freemanauto.com",
)
BAD_COMPANY_PREFIXES = (
    "source",
    "url",
    "contact",
    "contact us",
    "shop",
    "store",
    "products",
    "all products",
    "parts",
    "accessories",
    "brakes",
    "new arrivals",
    "best sellers",
    "terms",
    "terms of use",
    "privacy",
    "accessibility",
    "email sign-up",
    "home",
    "skip to content",
    "skip content",
    "menu",
    "navigation",
    "cart",
    "my account",
    "search",
    "user_daily_quota_exhausted",
    "daily_free_quota_exhausted",
    "anysearch_timeout",
    "anysearch_limited",
    "主页",
)
NOISE_EMAIL_DOMAINS = (
    "hmausa.com",
    "autoplay.co.nz",
    "shop.app",
)
MODEL_REQUIREMENTS = {
    "santa_fe": {
        "model_terms": ("santa fe", "santafe", "santa-fe", "mx5", "xrt"),
        "must_not_domains": ("hyundai.com", "hyundaiusa.com", "hyundaicanada.com", "hmausa.com"),
        "must_not_text": ("honda of santa fe", "santa fe online store", "santa fe boutique", "the collective santa fe"),
    },
    "jetour_t2_g700": {
        "model_terms": ("jetour", "t2", "g700", "traveller", "traveler", "jietu"),
        "must_not_domains": ("jetour.com", "jetourglobal.com", "jetourauto.com", "cheryinternational.com"),
        "must_not_text": ("official jetour", "jetour global", "dealer locator"),
    },
    "byd_shark6": {
        "model_terms": ("byd shark", "shark 6", "shark6"),
        "must_not_domains": ("byd.com", "autoplay.co.nz"),
        "must_not_text": ("vehicle listing", "book a test drive"),
    },
    "vw_amarok": {
        "model_terms": ("amarok", "volkswagen amarok", "vw amarok"),
        "must_not_domains": ("volkswagen.com", "vw.com"),
        "must_not_text": ("drivergear", "company store"),
    },
    "ranger_t9": {
        "model_terms": ("ranger", "ford ranger", "ranger t9", "next-gen ranger"),
        "must_not_domains": ("polaris.com", "rangerready.net", "rangerstation.co"),
        "must_not_text": ("polaris ranger", "ranger wear", "insect repellent"),
    },
}
MODEL_SEED_URLS = {
    "santa_fe": (
        "https://www.psashop.cz/en/santa-fe-mx5-prislusenstvi/",
        "https://carromats.ca/products/hyundai-santa-fe-gen4-floor-mats",
        "https://ahparts.com/Hyundai-SANTA-FE-replacement-parts",
        "https://www.am-autoparts.com/about/contact_us.html",
        "https://modernspare.com/product/2020-2025-hyundai-santa-fe-spare-tire-kit-options/",
        "https://www.audioledcar.com/",
        "https://cars-equipment.com/",
        "https://auovo.com/",
    ),
    "jetour_t2_g700": (
        "https://uoroffroad.com/brand/jetour/",
        "https://www.protuning.mu/",
        "https://cargadget.ae/",
        "https://desertleaders.com/",
        "https://desertgears.com/",
        "https://qatarpart.com/",
        "https://holdfast.co.za/",
        "https://www.thebakkieshop.co.za/",
        "https://www.autostyle.co.za/",
        "https://www.fbonline4x4accessories.co.za/",
        "https://slickridezinc.co.za/",
    ),
}
MODEL_SEARCH_SEEDS = {
    "santa_fe": (
        "Hyundai Santa Fe MX5 accessories contact email",
        "Hyundai Santa Fe 2024 roof rack store contact",
        "Hyundai Santa Fe 2024 side steps accessories email",
        "Hyundai Santa Fe MX5 cross bars shop email",
        "Hyundai Santa Fe aftermarket accessories contact",
    ),
    "jetour_t2_g700": (
        "site:.ae Jetour T2 accessories email",
        "site:.za Jetour T2 4x4 accessories email",
        "site:.mu Jetour T2 accessories contact",
        "site:.my Jetour T2 offroad accessories email",
        "site:.pk Jetour T2 accessories contact",
        "Jetour Traveller accessories shop email",
        "Jetour T2 roof rack side steps contact",
        "Jetour T2 offroad accessories dealer email",
        "Jetour G700 accessories dealer email",
    ),
}
ANYSEARCH_QUOTA_TERMS = (
    "user_daily_quota_exhausted",
    "daily_free_quota_exhausted",
    "quota exhausted",
    "rate limit",
    "too many requests",
    "429",
)
AUTO_REPLY_TERMS = (
    "automated reply",
    "automatic reply",
    "auto reply",
    "auto-reply",
    "do not reply",
    "no reply",
    "noreply",
    "no-reply",
    "request received",
    "we have received your message",
    "we've received your message",
    "your request has been received",
    "ticket was created",
    "support ticket",
    "how would you rate",
    "satisfaction",
    "please rate",
    "customer service survey",
    "view in browser",
    "unsubscribe",
    "price drop",
    "newsletter",
    "advertisement",
    "退订",
    "投诉",
    "自动回复",
    "系统邮件",
    "已收到您的邮件",
)
STRONG_INTEREST_PATTERNS = (
    r"\bsend me (?:a )?(?:catalog|catalogue|price|pricing|quote|quotation)",
    r"\bplease send (?:us |me )?(?:your )?(?:catalog|catalogue|price|pricing|quote|quotation|more information)",
    r"\bcatalog(?:ue)? and pric(?:e|ing)",
    r"\bprice list\b",
    r"\bpurchase prices?\b",
    r"\bhow much\b",
    r"\bminimum order\b",
    r"\bmoq\b",
    r"\bdelivery (?:time|times|cost|costs|fee|fees)",
    r"\blead time\b",
    r"\bwe (?:are|would be|might be) interested\b",
    r"\b(?:i am|i'm|im|we are|we're) interested(?: in)?\b",
    r"\binterested in (?:your|this|these|the|ranger|byd|amarok|santa|jetour|g700|product|products)\b",
    r"\bthat could indeed be quite interesting\b",
    r"\byour products are very interesting\b",
    r"\bwe can integrate your catalog",
    r"\binclude them on our website\b",
    r"\bwant (?:more )?(?:info|information)\b",
    r"\bwant to buy\b",
    r"\bwant buy\b",
    r"\bbuy something\b",
    r"\b(?:i|we) (?:want|need|would like|plan|hope) to (?:buy|purchase|order)\b",
    r"\binterested in (?:buying|purchasing|ordering)\b",
    r"\bwhere can i buy\b",
    r"\bhow (?:can|do) i (?:buy|purchase|order)\b",
    r"\bhow to (?:buy|purchase|order)\b",
    r"\bi need (?:one|two|[0-9]+|a set|sets?|pcs?|pieces?|units?|this|these|it)\b",
    r"\b(?:place|make) an order\b",
    r"\bready to (?:buy|purchase|order)\b",
    r"\bpurchase intent\b",
    r"\brfq\b",
    r"\brequest for (?:a )?(?:quote|quotation|pricing|price|proposal)\b",
    r"\b(?:quote|quotation|pricing|price list|price sheet|price offer)\b",
    r"\b(?:best|final|wholesale|dealer|distributor|retail) price\b",
    r"\bwhat(?:'s| is) (?:the )?(?:price|cost|moq|minimum order)\b",
    r"\bcould you (?:quote|send|provide|share|give)\b",
    r"\bcan you (?:quote|send|provide|share|give)\b",
    r"\b(?:send|share|provide|give) (?:(?:me|us)\s+)?(?:your )?(?:catalogue?|brochure|price list|pricing|quote|quotation|invoice|pi)\b",
    r"\b(?:send|share|provide|give) (?:(?:me|us)\s+)?(?:more )?(?:details|product details|specs|specifications)\b",
    r"\b(?:more|further) (?:details|product details|specs|specifications)\b",
    r"\b(?:catalogue?|brochure|product list|price list|excel file|price sheet)\b",
    r"\b(?:fitment|dimensions|installation|install guide)\b",
    r"\b(?:buy|purchase|order)(?:ing)? (?:one|some|a|an|the|your|this|these|[0-9]+|sets?|pieces?|pcs?|units?|sample|samples|body ?kit|accessor(?:y|ies)|product|products)\b",
    r"\b(?:place|make|start) (?:an? )?order\b",
    r"\b(?:trial|sample|bulk|wholesale|first|initial) order\b",
    r"\border (?:quantity|qty|details|now|today)\b",
    r"\b(?:qty|quantity) (?:of|for|required|needed|available)\b",
    r"\bminimum order quantity\b",
    r"\bminimum quantity\b",
    r"\bsample(?:s)? (?:available|cost|price|order|request)\b",
    r"\b(?:send|provide|ship) (?:a )?sample\b",
    r"\btest sample\b",
    r"\b(?:lead time|delivery time|production time|shipping time|dispatch time|turnaround time)\b",
    r"\b(?:shipping|shipment|freight|delivery) (?:cost|fee|fees|price|quote|time|to)\b",
    r"\bdo you ship\b",
    r"\bship to\b",
    r"\b(?:available|availability|in stock|stock available|ready stock)\b",
    r"\bcan you supply\b",
    r"\bare you able to supply\b",
    r"\blooking for (?:a )?(?:supplier|manufacturer|factory)\b",
    r"\b(?:payment terms?|payment method|bank transfer|paypal|credit card|proforma invoice|commercial invoice)\b",
    r"\bplease (?:send|issue) (?:a )?(?:pi|proforma invoice|invoice)\b",
    r"\bcan you (?:send|issue) (?:a )?(?:pi|proforma invoice|invoice)\b",
    r"\bpi for\b",
    r"\b(?:incoterms?|fob|cif|exw|dap|ddp)\b",
    r"\b(?:dealer|distributor|reseller|wholesale|wholesaler|retailer|dropship|dropshipping)\b",
    r"\bbecome (?:your )?(?:dealer|distributor|reseller)\b",
    r"\bdistribute your products\b",
    r"\bfor resale\b",
    r"\b(?:importer|importing|import your products)\b",
    r"\b(?:contact|call|whatsapp|message) (?:me|us)\b",
    r"\bmore info\b",
    r"\bmore information\b",
    r"\bfurther information\b",
    r"\bplease send\b",
    r"\bwebsite where i can have a look\b",
    r"\bis your .* compatible\b",
    r"\bcompatible with\b",
    r"\bcatalogues with excel files\b",
    r"\bgoogle drive links?\b",
    r"\binteresados\b",
    r"\binteresado\b",
    r"\bm[aá]s informaci[oó]n\b",
    r"\bprecios?\b",
    r"\bcotizaci[oó]n\b",
    r"\bpresupuesto\b",
    r"\blista de precios\b",
    r"\bquiero comprar\b",
    r"\bcomprar\b",
    r"\bhacer (?:un )?pedido\b",
    r"\bpedido\b",
    r"\bmuestras?\b",
    r"\bplazo de entrega\b",
    r"\benv[ií]o\b",
    r"\bdistribuidor(?:es)?\b",
    r"\bmayorista\b",
    r"\btiempos de entrega\b",
    r"\bcota[cç][aã]o\b",
    r"\bor[cç]amento\b",
    r"\bpre[cç]os?\b",
    r"\bcomprar\b",
    r"\bpedido\b",
    r"\bamostras?\b",
    r"\bprazo de entrega\b",
    r"\bfrete\b",
    r"\bdistribuidor(?:es)?\b",
    r"\batacado\b",
    r"\bdevis\b",
    r"\bprix\b",
    r"\btarifs?\b",
    r"\bacheter\b",
    r"\bcommander\b",
    r"\bcommande\b",
    r"\b[eé]chantillon\b",
    r"\bd[eé]lai de livraison\b",
    r"\bfrais de livraison\b",
    r"\bdistributeur\b",
    r"\bgrossiste\b",
    r"\bangebot\b",
    r"\bpreis(?:e|liste)?\b",
    r"\bkaufen\b",
    r"\bbestellen\b",
    r"\bbestellung\b",
    r"\bmuster\b",
    r"\blieferzeit\b",
    r"\bversand\b",
    r"\bh[aä]ndler\b",
    r"\bgro[ßs]handel\b",
    r"\bpreventivo\b",
    r"\bprezz[oi]\b",
    r"\blistino prezzi\b",
    r"\bcomprare\b",
    r"\bacquistare\b",
    r"\bordine\b",
    r"\bcampione\b",
    r"\btempi di consegna\b",
    r"\bspedizione\b",
    r"\bdistributore\b",
    r"\bgrossista\b",
    r"سعر",
    r"عرض سعر",
    r"شراء",
    r"أريد شراء",
    r"طلب",
    r"عينة",
    r"الشحن",
    r"التوصيل",
    r"موزع",
    r"جملة",
    r"цена",
    r"прайс",
    r"купить",
    r"заказать",
    r"заказ",
    r"образец",
    r"доставка",
    r"срок поставки",
    r"оплата",
    r"дистрибьютор",
    r"оптом",
    r"\bharga\b",
    r"\bbeli\b",
    r"\bmembeli\b",
    r"\bpesan\b",
    r"\bongkir\b",
    r"\bpengiriman\b",
    r"\bgrosir\b",
    r"価格",
    r"見積",
    r"購入",
    r"注文",
    r"サンプル",
    r"納期",
    r"送料",
    r"代理店",
    r"卸",
    r"가격",
    r"견적",
    r"구매",
    r"주문",
    r"샘플",
    r"배송",
    r"납기",
    r"대리점",
    r"도매",
    r"报价",
    r"询价",
    r"询问",
    r"咨询",
    r"想了解",
    r"了解一下",
    r"价格",
    r"目录",
    r"样品",
    r"采购",
    r"采购意向",
    r"购买",
    r"想买",
    r"我要买",
    r"下单",
    r"订购",
    r"购买意向",
    r"批发",
    r"经销",
    r"代理",
    r"库存",
    r"现货",
    r"发货",
    r"运费",
    r"交期",
    r"付款",
    r"发票",
    r"形式发票",
    r"样品费",
    r"最低起订量",
    r"起订量",
    r"合作",
    r"请发",
    r"发我",
)
COUNTRY_BY_SUFFIX = {
    ".com.cn": "中国",
    ".cn": "中国",
    ".com.hk": "香港",
    ".hk": "香港",
    ".com.tw": "台湾",
    ".tw": "台湾",
    ".com.au": "澳大利亚",
    ".au": "澳大利亚",
    ".co.nz": "新西兰",
    ".nz": "新西兰",
    ".co.za": "南非",
    ".za": "南非",
    ".com.br": "巴西",
    ".br": "巴西",
    ".com.mx": "墨西哥",
    ".mx": "墨西哥",
    ".ca": "加拿大",
    ".us": "美国",
    ".co.uk": "英国",
    ".uk": "英国",
    ".cz": "捷克",
    ".sk": "斯洛伐克",
    ".mu": "毛里求斯",
    ".my": "马来西亚",
    ".sg": "新加坡",
    ".id": "印度尼西亚",
    ".ae": "阿联酋",
    ".qa": "卡塔尔",
    ".sa": "沙特阿拉伯",
    ".eg": "埃及",
    ".lb": "黎巴嫩",
    ".jm": "牙买加",
    ".ke": "肯尼亚",
    ".ng": "尼日利亚",
    ".pk": "巴基斯坦",
    ".de": "德国",
    ".fr": "法国",
    ".it": "意大利",
    ".es": "西班牙",
    ".nl": "荷兰",
    ".pl": "波兰",
    ".th": "泰国",
    ".ph": "菲律宾",
    ".tr": "土耳其",
    ".ru": "俄罗斯",
    ".za": "南非",
    ".com": "",
}

TASK_LOCK = threading.Lock()
CURRENT_TASK: dict[str, Any] | None = None
MAIL_MONITOR_THREAD: threading.Thread | None = None
MAIL_MONITOR_STOP = threading.Event()
ANYSEARCH_DISABLED_UNTIL = 0.0
ANYSEARCH_DISABLED_REASON = ""


def now_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def read_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8-sig"))
    except Exception:
        return default


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def append_log(message: str) -> None:
    RUNTIME_DIR.mkdir(parents=True, exist_ok=True)
    with LOG_PATH.open("a", encoding="utf-8") as handle:
        handle.write(f"[{now_iso()}] {message}\n")


def deep_merge(base: Any, override: Any) -> Any:
    if isinstance(base, dict) and isinstance(override, dict):
        merged = dict(base)
        for key, value in override.items():
            merged[key] = deep_merge(merged.get(key), value)
        return merged
    return override if override is not None else base


def load_config() -> dict[str, Any]:
    config = read_json(CONFIG_PATH, {})
    local = read_json(LOCAL_CONFIG_PATH, {})
    return deep_merge(config, local)


def project_path(value: Any) -> Path:
    text = normalize(value)
    path = Path(text)
    if not text:
        return path
    return path if path.is_absolute() else ROOT / path


def python_executable(config: dict[str, Any]) -> str:
    configured = normalize(config.get("python"))
    if configured:
        if "/" in configured or "\\" in configured:
            candidate = project_path(configured)
            if candidate.exists():
                return str(candidate)
            return sys.executable
        return configured
    bundled = ROOT / "python" / "python.exe"
    return str(bundled) if bundled.exists() else sys.executable


def normalize(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def compact(value: Any) -> str:
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", normalize(value).lower())


def host_from_url(url: str) -> str:
    text = normalize(url)
    if not text:
        return ""
    if "://" not in text:
        text = "https://" + text
    try:
        host = urlparse(text).netloc.lower()
    except Exception:
        return ""
    if host.startswith("www."):
        host = host[4:]
    return host.split(":")[0]


def main_domain(value: str) -> str:
    host = host_from_url(value) or normalize(value).lower()
    if "@" in host:
        host = host.rsplit("@", 1)[1]
    parts = [part for part in host.split(".") if part]
    if len(parts) <= 2:
        return host
    suffix2 = ".".join(parts[-2:])
    suffix3 = ".".join(parts[-3:])
    if suffix2 in {"co.uk", "com.au", "co.nz", "co.za", "com.br", "com.mx", "com.cn", "com.hk", "com.tw"}:
        return suffix3
    return ".".join(parts[-2:])


def infer_country(url: str, text: str = "") -> str:
    host = host_from_url(url)
    for suffix, country in sorted(COUNTRY_BY_SUFFIX.items(), key=lambda item: len(item[0]), reverse=True):
        if host.endswith(suffix) and country:
            return country
    merged = f"{url} {text}".lower()
    preserve_region_hints = [
        ("hong kong", "香港"),
        ("hk", "香港"),
        ("taiwan", "台湾"),
    ]
    for token, country in preserve_region_hints:
        if re.search(rf"\b{re.escape(token)}\b", merged):
            return country
    if re.search(r"\b(china|mainland china|prc)\b", merged) or re.search(
        r"(中国|大陆|中华人民共和国|广东|广州|深圳|上海|北京|江苏|浙江|山东|河北|福建|河南|湖北|湖南|四川|重庆|天津|安徽|江西|广西|辽宁)",
        text,
    ):
        own_company_terms = (
            "changzhou",
            "xinbei district",
            "menghe",
            "jiangsu province",
            "celeste",
            "artway",
            "factory-direct from china",
            "made in china",
            "based in changzhou",
        )
        if not any(term in merged for term in own_company_terms):
            return "中国"
    hints = [
        ("australia", "澳大利亚"),
        ("australian", "澳大利亚"),
        ("new zealand", "新西兰"),
        ("south africa", "南非"),
        ("united kingdom", "英国"),
        ("great britain", "英国"),
        ("britain", "英国"),
        ("england", "英国"),
        ("uk", "英国"),
        ("germany", "德国"),
        ("france", "法国"),
        ("italy", "意大利"),
        ("spain", "西班牙"),
        ("thailand", "泰国"),
        ("philippines", "菲律宾"),
        ("brazil", "巴西"),
        ("mexico", "墨西哥"),
        ("uae", "阿联酋"),
        ("dubai", "阿联酋"),
    ]
    for token, country in hints:
        if re.search(rf"\b{re.escape(token)}\b", merged):
            return country
    return ""


def is_valid_email(candidate: str) -> bool:
    value = candidate.strip().strip(".,;:()[]<>").lower()
    if not EMAIL_RE.fullmatch(value):
        return False
    if value.startswith(BAD_EMAIL_PREFIXES):
        return False
    local, domain = value.split("@", 1)
    if len(local) < 2 or len(local) > 64 or len(value) > 254:
        return False
    if local.isdigit():
        return False
    if local in {"example", "test", "user"} or domain in {"example.com", "test.com", "localhost"}:
        return False
    if domain.endswith((".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg")):
        return False
    return True


def decode_mime_header(value: str | None) -> str:
    if not value:
        return ""
    result = []
    for part, enc in decode_header(value):
        if isinstance(part, bytes):
            charset = (enc or "utf-8").lower()
            if charset in {"unknown-8bit", "x-unknown", "unknown"}:
                charset = "utf-8"
            try:
                result.append(part.decode(charset, errors="replace"))
            except LookupError:
                result.append(part.decode("utf-8", errors="replace"))
        else:
            result.append(part)
    return "".join(result)


def parse_local_datetime(value: Any) -> datetime | None:
    text = normalize(value)
    if not text:
        return None
    for parser in (
        datetime.fromisoformat,
        lambda item: datetime.strptime(item, "%Y-%m-%d %H:%M:%S"),
        lambda item: datetime.strptime(item, "%Y-%m-%d"),
    ):
        try:
            return parser(text)
        except ValueError:
            continue
    return None


def message_datetime(msg: email.message.Message) -> datetime | None:
    raw = msg.get("Date")
    if not raw:
        return None
    try:
        parsed = parsedate_to_datetime(raw)
    except Exception:
        return None
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone().replace(tzinfo=None)
    return parsed


def should_ignore_bounce(config: dict[str, Any], msg: email.message.Message) -> bool:
    ignore_before = parse_local_datetime(config.get("mail_monitor", {}).get("ignore_bounces_before"))
    msg_time = message_datetime(msg)
    return bool(ignore_before and msg_time and msg_time < ignore_before)


def model_workbook_path(config: dict[str, Any], model_key: str) -> Path:
    model = config["models"][model_key]
    folder = project_path(config["excel_root"]) / model["folder"]
    configured = model.get("workbook", "")
    direct = folder / configured if configured else Path("")
    if configured and direct.exists():
        return direct
    candidates = [
        path
        for path in folder.glob("*.xlsx")
        if not path.name.lower().endswith((".bak.xlsx",)) and ".bak-" not in path.name.lower()
    ]
    if not candidates:
        return direct
    if configured:
        stem = Path(configured).stem.lower()
        for path in candidates:
            if path.stem.lower() == stem:
                return path
    return sorted(candidates, key=lambda item: item.stat().st_mtime, reverse=True)[0]


def model_image_paths(config: dict[str, Any], model: dict[str, Any]) -> tuple[list[Path], list[str]]:
    values = (
        model.get("images")
        or model.get("first_images")
        or model.get("image_paths")
        or model.get("image_path")
        or model.get("image")
        or []
    )
    if isinstance(values, str):
        values = [values]
    base_dirs = [
        project_path(value)
        for value in [
            config.get("email_picture_dir", ""),
            config.get("image_root", ""),
            config.get("outreach_package", ""),
        ]
        if value
    ]
    resolved: list[Path] = []
    missing: list[str] = []
    for value in values:
        text = normalize(value)
        if not text:
            continue
        path = Path(text)
        candidates = [path] if path.is_absolute() else [base / path for base in base_dirs]
        found = next((candidate for candidate in candidates if candidate.exists()), None)
        if found:
            resolved.append(found)
        else:
            missing.append(text)
    return resolved, missing


def headers_for_sheet(sheet: Any) -> list[Any]:
    return [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]


def header_index(headers: list[Any], names: list[str]) -> int | None:
    normalized = [compact(value) for value in headers]
    wanted = [compact(value) for value in names if value]
    for index, header in enumerate(normalized):
        if header and header in wanted:
            return index + 1
    for index, header in enumerate(normalized):
        if header and any(want and want in header for want in wanted):
            return index + 1
    return None


def sheet_sample_text(sheet: Any, rows: int = 8, cols: int = 12) -> str:
    values = [sheet.title]
    for row in range(1, min(sheet.max_row, rows) + 1):
        for col in range(1, min(sheet.max_column, cols) + 1):
            value = normalize(sheet.cell(row, col).value)
            if value:
                values.append(value)
    return "\n".join(values)


def token_hits(text: str, tokens: tuple[str, ...]) -> int:
    compacted = compact(text)
    return sum(1 for token in tokens if compact(token) and compact(token) in compacted)


def sheet_has_email_header(sheet: Any) -> bool:
    return bool(header_index(headers_for_sheet(sheet), ["客户邮箱", "邮箱", "email", "mail"]))


def resolve_destination_sheet_name(workbook: Any, source_sheet: Any, configured: str, default_name: str, kind: str) -> str:
    if configured and configured in workbook.sheetnames:
        return configured

    best_name = ""
    best_score = -999
    for sheet in workbook.worksheets:
        if sheet.title == source_sheet.title:
            continue
        text = sheet_sample_text(sheet)
        name_text = sheet.title
        header_text = "\n".join(normalize(value) for value in headers_for_sheet(sheet) if normalize(value))
        invalid_hits = token_hits(text, INVALID_SHEET_TOKENS)
        follow_hits = token_hits(text, FOLLOWUP_SHEET_TOKENS)
        reply_header_hits = token_hits(header_text, FOLLOWUP_REPLY_TOKENS)
        removed_hits = token_hits(name_text, ("removed", "删除", "移除", "清理"))
        score = -100 if removed_hits else 0
        if kind == "invalid":
            score += token_hits(name_text, INVALID_SHEET_TOKENS) * 8 + invalid_hits * 3
            score -= follow_hits * 2
        else:
            if token_hits(name_text, INVALID_SHEET_TOKENS):
                score -= 100
            score += reply_header_hits * 12
            score += token_hits(name_text, FOLLOWUP_SHEET_TOKENS) * 8 + follow_hits * 3
            score -= invalid_hits * (2 if reply_header_hits else 10)
            if sheet_has_email_header(sheet):
                score += 2
            if sheet.max_row > 1:
                score += 1
        if score > best_score:
            best_score = score
            best_name = sheet.title

    if best_name and best_score > 0:
        return best_name
    return configured or default_name


def field_aliases_for_model(model: dict[str, Any]) -> dict[str, list[str]]:
    return {
        "company": [model.get("company_header", ""), "客户名", "客户", "公司", "company", "name"],
        "country": [model.get("country_header", ""), "客户国家", "国家", "country"],
        "email": [model.get("email_header", ""), "客户邮箱", "邮箱", "email", "mail"],
        "website": [model.get("website_header", ""), "客户网址", "网址", "网站", "website", "url"],
        "first_time": [model.get("first_time_header", ""), "一次跟进时间", "first sent", "first_time"],
        "second_time": [model.get("second_time_header", ""), "二次跟进时间", "second sent", "second_time"],
        "third_time": [model.get("third_time_header", ""), "三次跟进时间", "third sent", "third_time"],
    }


def open_model_sheet(config: dict[str, Any], model_key: str, data_only: bool = False) -> tuple[Path, Any, Any, list[Any], dict[str, int]]:
    model = config["models"][model_key]
    path = model_workbook_path(config, model_key)
    workbook = load_workbook(path, data_only=data_only)
    sheet_name = model.get("sheet", "Sheet1")
    sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    headers = headers_for_sheet(sheet)
    mapping = {
        "email": header_index(headers, [model.get("email_header", ""), "客户邮箱", "邮箱", "email", "mail"]),
        "company": header_index(headers, [model.get("company_header", ""), "客户名", "客户", "公司", "company", "name"]),
        "country": header_index(headers, [model.get("country_header", ""), "客户国家", "国家", "country"]),
        "website": header_index(headers, [model.get("website_header", ""), "客户网址", "网址", "网站", "website", "url"]),
        "first_time": header_index(headers, [model.get("first_time_header", ""), "一次跟进时间", "first sent"]),
    }
    return path, workbook, sheet, headers, {k: v for k, v in mapping.items() if v}


def read_model_rows(config: dict[str, Any], model_key: str) -> list[dict[str, Any]]:
    path, workbook, sheet, _headers, mapping = open_model_sheet(config, model_key, data_only=True)
    rows = []
    try:
        email_col = mapping.get("email")
        if not email_col:
            return []
        for row_number in range(2, sheet.max_row + 1):
            email_value = normalize(sheet.cell(row_number, email_col).value)
            if not email_value:
                continue
            row = {
                "row": row_number,
                "email": email_value,
                "company": normalize(sheet.cell(row_number, mapping.get("company", 0)).value) if mapping.get("company") else "",
                "country": normalize(sheet.cell(row_number, mapping.get("country", 0)).value) if mapping.get("country") else "",
                "website": normalize(sheet.cell(row_number, mapping.get("website", 0)).value) if mapping.get("website") else "",
                "first_time": normalize(sheet.cell(row_number, mapping.get("first_time", 0)).value) if mapping.get("first_time") else "",
            }
            rows.append(row)
    finally:
        workbook.close()
    return rows


def count_invalid_sheet(config: dict[str, Any], model_key: str) -> int:
    model = config["models"][model_key]
    path = model_workbook_path(config, model_key)
    if not path.exists():
        return 0
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        name = model.get("invalid_sheet", "邮箱失效")
        if name not in workbook.sheetnames:
            return 0
        sheet = workbook[name]
        return max(0, sheet.max_row - 1)
    finally:
        workbook.close()


def today_invalid_entries() -> list[dict[str, Any]]:
    data = read_json(DAILY_INVALID_PATH, {})
    today = date.today().isoformat()
    if isinstance(data, dict):
        entries = data.get(today, [])
    elif isinstance(data, list):
        entries = [item for item in data if item.get("date") == today]
    else:
        entries = []
    return [item for item in entries if isinstance(item, dict)]


def count_today_invalid(model_key: str) -> int:
    return sum(1 for item in today_invalid_entries() if item.get("model") == model_key)


def record_today_invalid(model_key: str, email_value: str, sender_profile: str, reason: str) -> None:
    today = date.today().isoformat()
    data = read_json(DAILY_INVALID_PATH, {})
    if not isinstance(data, dict):
        data = {}
    entries = data.get(today, [])
    if not isinstance(entries, list):
        entries = []
    signature = (model_key, email_value.lower())
    for item in entries:
        if (item.get("model"), normalize(item.get("email")).lower()) == signature:
            item.update({"sender": sender_profile, "reason": reason, "updated_at": now_iso()})
            break
    else:
        entries.append(
            {
                "date": today,
                "time": now_iso(),
                "model": model_key,
                "email": email_value.lower(),
                "sender": sender_profile,
                "reason": reason[:500],
            }
        )
    data = {today: entries}
    write_json(DAILY_INVALID_PATH, data)


def dashboard_status() -> dict[str, Any]:
    config = load_config()
    models = []
    for key, model in config.get("models", {}).items():
        try:
            rows = read_model_rows(config, key)
            sent = sum(1 for row in rows if row.get("first_time"))
            models.append(
                {
                    "key": key,
                    "label": model["label"],
                    "count": len(rows),
                    "sent": sent,
                    "unsent": len(rows) - sent,
                    "invalid": count_today_invalid(key),
                    "last_updated": datetime.fromtimestamp(model_workbook_path(config, key).stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
                    if model_workbook_path(config, key).exists()
                    else "",
                }
            )
        except Exception as exc:
            models.append({"key": key, "label": model.get("label", key), "error": str(exc), "count": 0, "sent": 0, "unsent": 0, "invalid": 0})
    return {
        "now": now_iso(),
        "models": models,
        "task": CURRENT_TASK,
        "intervention": read_interventions(),
        "mail_monitor": {"running": MAIL_MONITOR_THREAD is not None and MAIL_MONITOR_THREAD.is_alive()},
        "mail_accounts": mail_account_statuses(config),
        "senders": sender_options(config),
    }


def tail_lines(path: Path, limit: int) -> list[str]:
    if not path.exists():
        return []
    lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
    return lines[-limit:]


def sender_profiles(config: dict[str, Any]) -> dict[str, Any]:
    package = project_path(config.get("outreach_package", ""))
    path = package / "sender_profiles.local.json"
    return read_json(path, {})


def sender_options(config: dict[str, Any]) -> list[dict[str, str]]:
    options = []
    for key, profile in sender_profiles(config).items():
        options.append(
            {
                "key": key,
                "label": key,
                "user": normalize(profile.get("user")),
                "host": normalize(profile.get("host")),
            }
        )
    return options


def resolve_mail_account(config: dict[str, Any], account: dict[str, Any]) -> dict[str, Any]:
    profiles = sender_profiles(config)
    profile = profiles.get(account.get("sender_profile"), {})
    folders = account.get("folders") or config.get("mail_monitor", {}).get("default_folders", ["INBOX"])
    return {
        "key": account.get("key") or account.get("label") or "mail",
        "label": account.get("label") or account.get("key") or "Mail",
        "imap_host": normalize(account.get("imap_host")),
        "imap_port": int(account.get("imap_port", 993)),
        "imap_user": normalize(account.get("imap_user")) or normalize(profile.get("user")),
        "imap_password": normalize(account.get("imap_password")) or normalize(profile.get("password")),
        "folders": folders,
    }


def mail_account_statuses(config: dict[str, Any]) -> list[dict[str, Any]]:
    saved = read_json(MAIL_STATUS_PATH, {})
    saved_messages = read_json(MAIL_MESSAGES_PATH, {})
    statuses = []
    monitor_running = MAIL_MONITOR_THREAD is not None and MAIL_MONITOR_THREAD.is_alive()
    for account in config.get("mail_accounts", []):
        resolved = resolve_mail_account(config, account)
        item = {
            "key": resolved["key"],
            "label": resolved["label"],
            "user": resolved["imap_user"],
            "status": "未配置" if not (resolved["imap_host"] and resolved["imap_user"] and resolved["imap_password"] and resolved["imap_password"] != "replace-me") else "等待检查",
            "last_checked": "",
            "checked": 0,
            "bounces": 0,
            "interested": 0,
            "folders": [],
            "error": "",
        }
        cached = saved.get(resolved["key"], {})
        item.update(cached)
        item["label"] = resolved["label"]
        item["user"] = resolved["imap_user"]
        if item.get("status") == "未配置" and resolved["imap_host"] and resolved["imap_user"] and resolved["imap_password"] and resolved["imap_password"] != "replace-me":
            item["status"] = "等待检查"
            item["error"] = ""
        if monitor_running and item.get("status") == "等待检查" and not item.get("last_checked"):
            item["status"] = "检查中"
        messages = saved_messages.get(resolved["key"], []) if isinstance(saved_messages, dict) else []
        if isinstance(messages, list):
            item["messages"] = messages[:5]
            item["new_count"] = sum(1 for message in messages if not message.get("read_at"))
        else:
            item["messages"] = []
            item["new_count"] = 0
        statuses.append(item)
    return statuses


def anysearch_cmd(config: dict[str, Any]) -> list[str]:
    raw = config.get("anysearch_command", "")
    if not raw:
        return []
    if raw.lower().startswith("python "):
        parts = ["python"] + raw.split(" ", 1)[1].split()
    else:
        parts = raw.split()
    if parts and parts[0].lower() in {"python", "python3"}:
        parts[0] = python_executable(config)
    for index, part in enumerate(parts[1:], start=1):
        if "/" not in part and "\\" not in part:
            continue
        candidate = project_path(part)
        if candidate.exists():
            parts[index] = str(candidate)
    return parts


def run_anysearch(config: dict[str, Any], args: list[str], timeout: int = 60) -> str:
    global ANYSEARCH_DISABLED_UNTIL, ANYSEARCH_DISABLED_REASON
    if time.time() < ANYSEARCH_DISABLED_UNTIL:
        return ANYSEARCH_DISABLED_REASON
    cmd = anysearch_cmd(config) + args
    if not cmd:
        return ""
    env = os.environ.copy()
    api_key = normalize(config.get("anysearch_api_key"))
    if api_key:
        env["ANYSEARCH_API_KEY"] = api_key
    try:
        proc = subprocess.run(
            cmd,
            cwd=str(ROOT),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=timeout,
            env=env,
        )
    except subprocess.TimeoutExpired:
        ANYSEARCH_DISABLED_UNTIL = time.time() + 600
        ANYSEARCH_DISABLED_REASON = "anysearch_timeout"
        append_log(f"AnySearch 超时，10 分钟内切换备用采集方式: {' '.join(args)}")
        return ANYSEARCH_DISABLED_REASON
    combined = "\n".join(part for part in [proc.stdout, proc.stderr] if part)
    if proc.returncode != 0:
        append_log(f"AnySearch failed: {' '.join(args)} :: {proc.stderr.strip()}")
    if anysearch_limited(combined):
        ANYSEARCH_DISABLED_UNTIL = time.time() + 3600
        ANYSEARCH_DISABLED_REASON = combined[:2000] or "anysearch_limited"
        append_log("AnySearch 额度不可用，切换备用采集方式")
    return combined


def anysearch_limited(output: str) -> bool:
    lower = output.lower()
    return any(term in lower for term in ANYSEARCH_QUOTA_TERMS)


def http_fetch(url: str, timeout: int = 10) -> str:
    request = Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/125.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        },
    )
    with urlopen(request, timeout=timeout) as response:
        raw = response.read(2_000_000)
        content_type = response.headers.get("Content-Type", "")
    charset_match = re.search(r"charset=([\w.\-]+)", content_type, re.I)
    charset = charset_match.group(1) if charset_match else "utf-8"
    return raw.decode(charset, errors="replace")


def html_to_text(html: str) -> str:
    text = re.sub(r"(?is)<(script|style|noscript).*?</\1>", " ", html)
    text = re.sub(r"(?is)<br\s*/?>", "\n", text)
    text = re.sub(r"(?is)</(p|div|li|tr|h[1-6])>", "\n", text)
    text = re.sub(r"(?is)<[^>]+>", " ", text)
    return re.sub(r"[ \t\r\f\v]+", " ", unescape(text))


def clean_result_url(raw: str) -> str:
    value = unescape(unquote(raw)).strip().strip(".,)\"'")
    if value.startswith("//"):
        value = "https:" + value
    parsed = urlparse(value)
    host = parsed.netloc.lower()
    if "duckduckgo.com" in host:
        target = parse_qs(parsed.query).get("uddg", [""])[0]
        if target:
            value = unquote(target)
            parsed = urlparse(value)
            host = parsed.netloc.lower()
    if "bing.com" in host and parsed.path.startswith("/ck/"):
        target = parse_qs(parsed.query).get("u", [""])[0]
        if target.startswith("a1"):
            payload = target[2:]
            payload += "=" * (-len(payload) % 4)
            try:
                value = base64.urlsafe_b64decode(payload).decode("utf-8", errors="replace")
            except Exception:
                pass
        elif target.startswith("http"):
            value = unquote(target)
    return value


def blocked_host(host: str) -> bool:
    normalized = host.lower().removeprefix("www.")
    if normalized in BAD_HOSTS:
        return True
    return any(normalized == suffix or normalized.endswith("." + suffix) for suffix in BAD_HOST_SUFFIXES)


def blocked_url(url: str) -> bool:
    lower = url.lower()
    host = host_from_url(url)
    if not host or blocked_host(host):
        return True
    return any(fragment in lower for fragment in BAD_URL_FRAGMENTS)


def parse_search_results(output: str) -> list[str]:
    urls = []
    for line in output.splitlines():
        if "**URL**:" in line:
            url = line.split("**URL**:", 1)[1].strip()
            urls.append(url)
    urls.extend(URL_RE.findall(output))
    clean = []
    for url in urls:
        url = clean_result_url(url)
        host = host_from_url(url)
        if not host:
            continue
        if blocked_url(url):
            continue
        if url not in clean:
            clean.append(url)
    return clean


def query_relevant_url(query: str, url: str, context: str = "") -> bool:
    source = re.sub(r"[^a-z0-9]+", " ", f"{url} {context}".lower())
    vehicle_tokens = (
        "ranger",
        "ford ranger",
        "byd shark",
        "shark 6",
        "amarok",
        "volkswagen amarok",
        "santa fe",
        "santafe",
        "mx5",
        "jetour",
        "g700",
        "traveller",
        "traveler",
    )
    accessory_tokens = ACCESSORY_TERMS + ("4x4", "offroad", "off road", "aftermarket")
    return any(token in source for token in vehicle_tokens) and any(token.replace("-", " ") in source for token in accessory_tokens)


def parse_fallback_search_results(query: str, output: str) -> list[str]:
    candidates: list[tuple[str, str]] = []
    for item in re.findall(r"(?is)<item\b.*?</item>", output):
        link_match = re.search(r"(?is)<link>\s*(.*?)\s*</link>", item)
        if link_match:
            candidates.append((link_match.group(1), html_to_text(item)))
    for href in re.findall(r'(?is)<a[^>]+href=["\']([^"\']+)["\'][^>]*>(.*?)</a>', output):
        url, label = href
        if "/ck/a" in url or "uddg=" in url or url.startswith("http"):
            candidates.append((url, html_to_text(label)))
    clean: list[str] = []
    for raw_url, context in candidates:
        url = clean_result_url(raw_url)
        if blocked_url(url):
            continue
        if not query_relevant_url(query, url, context):
            continue
        if url not in clean:
            clean.append(url)
    return clean


def fallback_search_urls(query: str, max_results: int = 12) -> list[str]:
    urls: list[str] = []
    search_urls = [
        f"https://www.bing.com/search?format=rss&q={quote_plus(query)}",
        f"https://www.bing.com/search?q={quote_plus(query)}&count={max_results}",
        f"https://duckduckgo.com/html/?q={quote_plus(query)}",
    ]
    for search_url in search_urls:
        if len(urls) >= max_results:
            break
        try:
            output = http_fetch(search_url, timeout=8)
        except Exception as exc:
            append_log(f"备用搜索失败: {query} :: {exc}")
            continue
        for url in parse_fallback_search_results(query, output):
            if url not in urls:
                urls.append(url)
            if len(urls) >= max_results:
                break
    return urls


def search_candidate_urls(config: dict[str, Any], query: str, max_results: int = 12) -> list[str]:
    output = run_anysearch(config, ["search", query, "--max_results", str(max_results)], timeout=25)
    urls = parse_search_results(output)
    if anysearch_limited(output) or len(urls) < max_results:
        for url in fallback_search_urls(query, max_results=max_results):
            if url not in urls:
                urls.append(url)
            if len(urls) >= max_results:
                break
    return urls[:max_results]


def batch_search_candidate_urls(config: dict[str, Any], queries: list[str], max_results: int = 8) -> dict[str, list[str]]:
    if not queries:
        return {}
    payload = json.dumps([{"query": query, "max_results": max_results} for query in queries], ensure_ascii=False)
    output = run_anysearch(config, ["batch_search", "--queries", payload], timeout=35)
    urls = parse_search_results(output)
    result: dict[str, list[str]] = {query: [] for query in queries}
    if urls:
        for query in queries:
            result[query] = urls[:]
    if anysearch_limited(output) or not any(result.values()):
        for query in queries:
            result[query] = search_candidate_urls(config, query, max_results=max_results)
    return result


def extract_page(config: dict[str, Any], url: str, use_anysearch: bool = True) -> str:
    output = run_anysearch(config, ["extract", url], timeout=45) if use_anysearch else ""
    if output.strip() and not anysearch_limited(output) and output not in {"anysearch_timeout", "anysearch_limited"}:
        return output[:120000]
    try:
        return html_to_text(http_fetch(url, timeout=8))[:120000]
    except Exception as exc:
        append_log(f"备用页面抓取失败: {url} :: {exc}")
        return ""


def related_urls(url: str, model: dict[str, Any] | None = None) -> list[str]:
    parsed = urlparse(url if "://" in url else "https://" + url)
    if not parsed.netloc:
        return []
    base = f"{parsed.scheme or 'https'}://{parsed.netloc}"
    model_terms = [term for term in (model or {}).get("vehicle_terms", []) if len(term) >= 4][:3]
    paths = [
        parsed.path or "/",
        "/contact",
        "/contact-us",
        "/pages/contact-us",
        "/about-us",
        "/about",
        "/privacy-policy",
    ]
    for term in model_terms:
        slug = quote_plus(term)
        paths.extend(
            [
                f"/search?q={slug}",
                f"/?s={slug}",
                f"/collections/{slug.replace('+', '-')}",
                f"/product-category/{slug.replace('+', '-')}",
            ]
        )
    urls = []
    for path in paths:
        if not path.startswith("/"):
            path = "/" + path
        candidate = base + path
        if candidate not in urls:
            urls.append(candidate)
    return urls[:10]


def extract_candidate_bundle(config: dict[str, Any], model: dict[str, Any], url: str) -> tuple[str, dict[str, str]]:
    pieces: list[str] = []
    page_texts: dict[str, str] = {}
    empty_or_failed = 0
    saw_accessory_page = False
    for index, candidate in enumerate(related_urls(url, model)):
        host = host_from_url(candidate)
        if not host or blocked_host(host):
            continue
        text = extract_page(config, candidate, use_anysearch=(index == 0))
        if not text:
            empty_or_failed += 1
            if empty_or_failed >= 2:
                break
            continue
        lower = f"{candidate} {text}".lower()
        if any(term in lower for term in ACCESSORY_TERMS):
            saw_accessory_page = True
        if index == 0 and not saw_accessory_page and not any(term in lower for term in SHOP_INTENT_TERMS):
            return text[:120000], {candidate: text} if text else {}
        page_texts[candidate] = text
        pieces.append(f"\nSource URL: {candidate}\n{text}")
        joined_length = sum(len(piece) for piece in pieces)
        if joined_length > 120000 or (saw_accessory_page and index >= 1 and EMAIL_RE.search("\n".join(pieces))):
            break
    return "\n".join(pieces)[:160000], page_texts


def evaluate_candidate_url(
    config: dict[str, Any],
    model: dict[str, Any],
    url: str,
    existing_emails: set[str],
    existing_domains: set[str],
) -> tuple[dict[str, str] | None, str]:
    if blocked_url(url):
        return None, "网址被屏蔽"
    domain = main_domain(url)
    if domain in existing_domains:
        return None, "重复域名"
    text, page_texts = extract_candidate_bundle(config, model, url)
    if not page_qualifies(model, url, text):
        return None, "车型或配件证据不足"
    emails = [email_value.lower() for email_value in EMAIL_RE.findall(text)]
    selected_email = select_business_email(emails, existing_emails, existing_domains, host_from_url(url))
    if not selected_email:
        return None, "未找到可用业务邮箱"
    email_country = infer_country(f"https://{selected_email.rsplit('@', 1)[-1]}", "")
    lead = {
        "company": candidate_company_from_text(url, text),
        "country": infer_country(url, text) or email_country,
        "email": selected_email,
        "website": url,
        "evidence": next(iter(page_texts.keys()), url),
    }
    ok, reason = validate_lead(model, lead, text)
    if not ok:
        return None, reason
    return lead, ""


def candidate_company_from_text(url: str, text: str) -> str:
    title_match = re.search(r"(?im)^\s*(?:#\s*)?(.{3,90})\s*$", text)
    if title_match:
        title = clean_company_name(title_match.group(1), url)
        if title and not looks_like_product_title(title):
            return title
    for line in text.splitlines():
        cleaned = clean_company_name(line, url)
        if cleaned and not looks_like_product_title(cleaned):
            return cleaned
    host = host_from_url(url)
    return host.split(".")[0].replace("-", " ").title() if host else "Unknown"


def looks_like_product_title(value: str) -> bool:
    lower = value.lower()
    product_hits = sum(1 for term in ACCESSORY_TERMS if term in lower)
    vehicle_hits = sum(1 for term in ("hyundai", "santa fe", "jetour", "ranger", "amarok", "byd", "shark") if term in lower)
    product_shape = any(token in lower for token in (" for ", " fits ", "compatible", "2024", "2025", "2019", "2020", "mx5"))
    return product_hits >= 1 and (vehicle_hits >= 1 or product_shape)


def clean_company_name(value: str, url: str = "") -> str:
    text = unescape(normalize(value)).strip(" #*-|")
    text = re.sub(r"(?i)^source\*\*\s*:?\s*", "", text).strip()
    text = re.sub(r"(?i)^\[?source\]?\s*:?\s*", "", text).strip()
    text = re.sub(r"(?i)^title\s*:?\s*", "", text).strip()
    text = re.sub(r"(?i)^url\s*:?\s*", "", text).strip()
    if not text or text.startswith(("http://", "https://")):
        return ""
    lower = text.lower()
    if any(lower == prefix or lower.startswith(prefix + " ") for prefix in BAD_COMPANY_PREFIXES):
        return ""
    if "@" in text or len(text) < 3 or len(text) > 90:
        return ""
    if len(re.sub(r"[^a-zA-Z\u4e00-\u9fff]", "", text)) < 3:
        return ""
    host = host_from_url(url)
    if host and text.lower() in {"home", host, "www." + host}:
        return ""
    return text[:90]


def model_requirement(model: dict[str, Any]) -> dict[str, tuple[str, ...]]:
    direct_key = normalize(model.get("_key"))
    if direct_key:
        return MODEL_REQUIREMENTS.get(direct_key, {})
    for key, item in load_config().get("models", {}).items():
        if item.get("label") == model.get("label"):
            return MODEL_REQUIREMENTS.get(key, {})
    return {}


def page_qualifies(model: dict[str, Any], url: str, text: str) -> bool:
    merged = f"{url} {text}".lower()
    host = host_from_url(url)
    if blocked_url(url):
        return False
    req = model_requirement(model)
    if any(host == domain or host.endswith("." + domain) for domain in req.get("must_not_domains", ())):
        return False
    if any(term in merged for term in req.get("must_not_text", ())):
        return False
    vehicle_terms = [term.lower() for term in model.get("vehicle_terms", [])]
    vehicle_terms.extend(req.get("model_terms", ()))
    has_vehicle = any(term in merged for term in vehicle_terms)
    has_accessory = any(term in merged for term in ACCESSORY_TERMS)
    has_shop_intent = any(term in merged for term in SHOP_INTENT_TERMS)
    return has_vehicle and has_accessory and has_shop_intent


def select_business_email(emails: list[str], existing_emails: set[str], existing_domains: set[str], page_host: str) -> str:
    scored: list[tuple[int, str]] = []
    for email_value in emails:
        value = email_value.lower()
        if not is_valid_email(value):
            continue
        email_domain = main_domain(value)
        if value in existing_emails or email_domain in existing_domains:
            continue
        if any(email_domain == domain or email_domain.endswith("." + domain) for domain in NOISE_EMAIL_DOMAINS):
            continue
        score = 0
        local = value.split("@", 1)[0]
        if email_domain == main_domain(page_host):
            score += 8
        if local in {"info", "sales", "contact", "hello", "office", "support", "parts"}:
            score += 5
        if local.startswith(("sales", "info", "contact")):
            score += 2
        if local in {"accessibility", "privacy", "legal", "careers", "jobs", "newsletter"}:
            score -= 10
        scored.append((score, value))
    if not scored:
        return ""
    scored.sort(key=lambda item: (-item[0], item[1]))
    return scored[0][1]


def validate_lead(model: dict[str, Any], lead: dict[str, str], text: str, require_page_evidence: bool = True) -> tuple[bool, str]:
    if not lead.get("company") or lead["company"].lower().startswith("source"):
        return False, "公司名无效"
    if not lead.get("email") or not is_valid_email(lead["email"]):
        return False, "邮箱无效"
    if not lead.get("website") or blocked_url(lead["website"]):
        return False, "网址无效"
    if require_page_evidence and not page_qualifies(model, lead["website"], text):
        return False, "车型或配件证据不足"
    country = lead.get("country", "")
    if country and country in set(model.get("exclude_countries", [])):
        return False, f"排除国家 {country}"
    email_domain = lead["email"].rsplit("@", 1)[-1]
    email_country = infer_country(f"https://{email_domain}", "")
    if email_country and email_country in set(model.get("exclude_countries", [])):
        return False, f"邮箱域名属于排除国家 {email_country}"
    return True, ""


def build_dedupe_sets(config: dict[str, Any], model_key: str) -> tuple[set[str], set[str]]:
    rows = read_model_rows(config, model_key)
    emails = {row["email"].strip().lower() for row in rows if row.get("email")}
    domains = {main_domain(row["website"]) for row in rows if row.get("website")}
    domains |= {main_domain(row["email"]) for row in rows if row.get("email")}
    domains.discard("")
    return emails, domains


def model_with_key(config: dict[str, Any], model_key: str) -> dict[str, Any]:
    return {**config["models"][model_key], "_key": model_key}


def lead_search_queries(model: dict[str, Any]) -> list[str]:
    queries: list[str] = []

    def add(value: str) -> None:
        text = normalize(value)
        if text and text.lower() not in {item.lower() for item in queries}:
            queries.append(text)

    for query in model.get("search_queries", []):
        add(query)

    for query in MODEL_SEARCH_SEEDS.get(normalize(model.get("_key")), ()):
        add(query)

    vehicle_terms = [term for term in model.get("vehicle_terms", []) if len(term) >= 4][:4]
    product_items = [term for term in model.get("message", {}).get("product_items", []) if len(term) >= 4][:4]
    for vehicle in vehicle_terms:
        for suffix in LEAD_SEARCH_SUFFIXES:
            add(f"{vehicle} {suffix}")
        for product in product_items:
            add(f"{vehicle} {product} dealer email")
            add(f"{vehicle} {product} supplier contact")

    return queries[:60]


def collect_for_model(config: dict[str, Any], model_key: str, target_count: int) -> dict[str, Any]:
    model = model_with_key(config, model_key)
    append_log(f"开始获客: {model['label']} 目标 {target_count}")
    existing_emails, existing_domains = build_dedupe_sets(config, model_key)
    found: list[dict[str, str]] = []
    rejected: list[dict[str, str]] = []
    seen_urls: set[str] = set()

    def try_url(url: str) -> None:
        if len(found) >= target_count or url in seen_urls:
            return
        seen_urls.add(url)
        lead, reason = evaluate_candidate_url(config, model, url, existing_emails, existing_domains)
        if not lead:
            append_log(f"跳过 {model['label']} {url}: {reason}")
            if len(rejected) < 40:
                rejected.append({"url": url, "reason": reason})
            return
        found.append(lead)
        existing_emails.add(lead["email"])
        existing_domains.add(main_domain(lead["website"]))
        existing_domains.add(main_domain(lead["email"]))
        append_log(f"{model['label']} 新客户 {len(found)}/{target_count}: {lead['company']} {lead['email']}")

    for seed_url in MODEL_SEED_URLS.get(model_key, ()):
        try_url(seed_url)
        if len(found) >= target_count:
            break

    queries = lead_search_queries(model)
    query_batches = [queries[index : index + 4] for index in range(0, len(queries), 4)]
    for query_batch in query_batches:
        if len(found) >= target_count:
            break
        batch_results = batch_search_candidate_urls(config, query_batch, max_results=10)
        batch_urls: list[str] = []
        for values in batch_results.values():
            for value in values:
                if value not in batch_urls:
                    batch_urls.append(value)
        urls = batch_urls
        if not urls:
            for query_text in query_batch:
                for value in search_candidate_urls(config, query_text, max_results=10):
                    if value not in urls:
                        urls.append(value)
        for url in urls:
            if len(found) >= target_count:
                break
            try_url(url)
    write_result = append_leads(config, model_key, found)
    written = write_result["count"]
    if written < target_count:
        append_log(f"完成获客: {model['label']} 写入 {written}/{target_count}，不足原因样本: {rejected[:8]}")
    else:
        append_log(f"完成获客: {model['label']} 写入 {written}/{target_count}")
    return {
        "model": model_key,
        "label": model["label"],
        "found": len(found),
        "written": write_result["count"],
        "written_emails": write_result["emails"],
        "leads": found,
        "shortfall": max(0, target_count - written),
        "rejected_sample": rejected[:20],
    }


def preview_collect_for_model(config: dict[str, Any], model_key: str, target_count: int = 10, query_limit: int = 8) -> dict[str, Any]:
    model = model_with_key(config, model_key)
    existing_emails, existing_domains = build_dedupe_sets(config, model_key)
    found: list[dict[str, str]] = []
    rejected: list[dict[str, str]] = []
    seen_urls: set[str] = set()

    def try_url(url: str) -> None:
        if len(found) >= target_count or url in seen_urls:
            return
        seen_urls.add(url)
        lead, reason = evaluate_candidate_url(config, model, url, existing_emails, existing_domains)
        if lead:
            found.append(lead)
            existing_emails.add(lead["email"])
            existing_domains.add(main_domain(lead["website"]))
            existing_domains.add(main_domain(lead["email"]))
        elif len(rejected) < 40:
            rejected.append({"url": url, "reason": reason})

    for seed_url in MODEL_SEED_URLS.get(model_key, ()):
        try_url(seed_url)
        if len(found) >= target_count:
            break

    queries = lead_search_queries(model)[:query_limit]
    for query in queries:
        if len(found) >= target_count:
            break
        for url in search_candidate_urls(config, query, max_results=8):
            if len(found) >= target_count:
                break
            try_url(url)
    return {
        "model": model_key,
        "label": model["label"],
        "target": target_count,
        "queries_checked": len(queries),
        "found": found,
        "rejected_sample": rejected[:20],
    }


def append_leads(config: dict[str, Any], model_key: str, leads: list[dict[str, str]]) -> dict[str, Any]:
    if not leads:
        return {"count": 0, "emails": []}
    model = model_with_key(config, model_key)
    path, workbook, sheet, headers, mapping = open_model_sheet(config, model_key, data_only=False)
    try:
        if not mapping.get("email"):
            raise RuntimeError("找不到邮箱列")
        required = ("company", "country", "email", "website")
        missing = [field for field in required if not mapping.get(field)]
        if missing:
            raise RuntimeError(f"找不到必要列: {', '.join(missing)}")
        backup = path.with_name(f"{path.stem}.bak-{datetime.now().strftime('%Y%m%d-%H%M%S')}{path.suffix}")
        shutil.copy2(path, backup)
        existing_emails = set()
        existing_domains = set()
        email_col = mapping.get("email")
        website_col = mapping.get("website")
        for row_number in range(2, sheet.max_row + 1):
            email_value = normalize(sheet.cell(row_number, email_col).value).lower()
            website_value = normalize(sheet.cell(row_number, website_col).value) if website_col else ""
            if email_value:
                existing_emails.add(email_value)
                existing_domains.add(main_domain(email_value))
            if website_value:
                existing_domains.add(main_domain(website_value))
        written = 0
        written_emails: list[str] = []
        for lead in leads:
            lead = {
                **lead,
                "company": clean_company_name(lead.get("company", ""), lead.get("website", "")),
                "email": normalize(lead.get("email")).lower(),
                "website": normalize(lead.get("website")),
                "country": normalize(lead.get("country")),
            }
            if not lead["country"]:
                lead["country"] = infer_country(f"https://{lead['email'].rsplit('@', 1)[-1]}", "")
            ok, reason = validate_lead(model, lead, f"{lead['company']} {lead['website']}", require_page_evidence=False)
            if not ok:
                append_log(f"写入前跳过 {model['label']} {lead.get('website', '')}: {reason}")
                continue
            email_value = lead["email"]
            website_domain = main_domain(lead["website"])
            email_domain = main_domain(email_value)
            if email_value in existing_emails or website_domain in existing_domains or email_domain in existing_domains:
                continue
            row = sheet.max_row + 1
            sheet.cell(row, mapping["company"]).value = lead["company"]
            sheet.cell(row, mapping["country"]).value = lead["country"]
            sheet.cell(row, mapping["email"]).value = lead["email"]
            sheet.cell(row, mapping["website"]).value = lead["website"]
            existing_emails.add(email_value)
            existing_domains.add(website_domain)
            existing_domains.add(email_domain)
            written += 1
            written_emails.append(email_value)
        workbook.save(path)
        return {"count": written, "emails": written_emails}
    finally:
        workbook.close()


def cleanup_blocked_leads(config: dict[str, Any], model_key: str) -> dict[str, Any]:
    model = model_with_key(config, model_key)
    path, workbook, sheet, _headers, mapping = open_model_sheet(config, model_key, data_only=False)
    try:
        email_col = mapping.get("email")
        website_col = mapping.get("website")
        if not email_col or not website_col:
            return {"model": model_key, "removed": 0, "rows": []}
        backup = path.with_name(f"{path.stem}.cleanup-bak-{datetime.now().strftime('%Y%m%d-%H%M%S')}{path.suffix}")
        rows_to_delete: list[tuple[int, str, str]] = []
        for row_number in range(2, sheet.max_row + 1):
            email_value = normalize(sheet.cell(row_number, email_col).value).lower()
            website = normalize(sheet.cell(row_number, website_col).value)
            if not email_value and not website:
                continue
            reason = ""
            if website and blocked_host(host_from_url(website)):
                reason = "屏蔽主机"
            elif email_value and is_valid_email(email_value):
                email_country = infer_country(f"https://{email_value.rsplit('@', 1)[-1]}", "")
                if email_country and email_country in set(model.get("exclude_countries", [])):
                    reason = f"排除国家邮箱 {email_country}"
            if reason:
                rows_to_delete.append((row_number, website or email_value, reason))
        if not rows_to_delete:
            return {"model": model_key, "removed": 0, "rows": []}
        shutil.copy2(path, backup)
        for row_number, _value, _reason in reversed(rows_to_delete):
            sheet.delete_rows(row_number, 1)
        workbook.save(path)
        append_log(f"清理旧获客噪声: {model['label']} 删除 {len(rows_to_delete)} 行 backup={backup}")
        return {
            "model": model_key,
            "removed": len(rows_to_delete),
            "backup": str(backup),
            "rows": [{"row": row, "value": value, "reason": reason} for row, value, reason in rows_to_delete],
        }
    finally:
        workbook.close()


def send_with_fallback_for_target(model_key: str, target_email: str, primary_profile: str | None, context: str) -> dict[str, Any]:
    config = load_config()
    primary = normalize(primary_profile) or normalize(config.get("sender_profile"))
    retry = normalize(config.get("retry_sender_profile"))
    attempts: list[dict[str, Any]] = []
    try:
        result = send_target_email_for_model(model_key, target_email, primary, log_label=context)
        attempts.append({"sender": primary, "ok": True})
        return {"ok": True, "result": result, "attempts": attempts}
    except Exception as first_exc:
        attempts.append({"sender": primary, "ok": False, "error": str(first_exc)})
        if not retry or retry == primary:
            moved = mark_final_invalid_email(config, model_key, target_email, primary, str(first_exc), "最终发送失败")
            return {"ok": False, "attempts": attempts, "moved_to_invalid": moved}
        try:
            result = send_target_email_for_model(model_key, target_email, retry, log_label=f"{context}备用发送")
            attempts.append({"sender": retry, "ok": True})
            return {"ok": True, "result": result, "attempts": attempts}
        except Exception as second_exc:
            attempts.append({"sender": retry, "ok": False, "error": str(second_exc)})
            moved = mark_final_invalid_email(config, model_key, target_email, retry, str(second_exc), "备用邮箱发送失败")
            return {"ok": False, "attempts": attempts, "moved_to_invalid": moved}


def daily_collect(limit_per_model: int) -> dict[str, Any]:
    config = load_config()
    results = []
    for key in config.get("models", {}):
        cleanup = cleanup_blocked_leads(config, key)
        result = collect_for_model(config, key, limit_per_model)
        result["cleanup"] = cleanup
        result["auto_sent"] = []
        results.append(result)
    return {"daily_limit": limit_per_model, "results": results}


def reset_model_send_state(model_key: str) -> dict[str, Any]:
    config = load_config()
    model = config["models"][model_key]
    path, workbook, sheet, _headers, mapping = open_model_sheet(config, model_key, data_only=False)
    try:
        first_col = mapping.get("first_time")
        email_col = mapping.get("email")
        if not first_col or not email_col:
            raise RuntimeError("找不到发送状态列或邮箱列")
        backup = path.with_name(f"{path.stem}.reset-bak-{datetime.now().strftime('%Y%m%d-%H%M%S')}{path.suffix}")
        shutil.copy2(path, backup)
        reset_count = 0
        for row_number in range(2, sheet.max_row + 1):
            if normalize(sheet.cell(row_number, email_col).value):
                sheet.cell(row_number, first_col).value = None
                reset_count += 1
        workbook.save(path)
        append_log(f"reset send state: {model['label']} rows={reset_count} backup={backup}")
        return {
            "model": model_key,
            "label": model["label"],
            "reset_rows": reset_count,
            "backup": str(backup),
        }
    finally:
        workbook.close()


def ensure_outreach_config(config: dict[str, Any], model_key: str, sender_profile: str | None = None) -> Path:
    model = config["models"][model_key]
    GENERATED_CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    package = project_path(config["outreach_package"])
    logo_file = ensure_fallback_logo()
    image_paths, missing_images = model_image_paths(config, model)
    if missing_images:
        raise RuntimeError(f"{model['label']} 邮件图片不存在: {', '.join(missing_images)}")
    if not image_paths:
        raise RuntimeError(f"{model['label']} 未配置邮件图片，已停止发送，避免发出无图邮件")
    cfg = {
        "campaign": {
            "name": f"{model_key}_first_touch",
            "mode": "first_touch_only",
            "sheet_name": model.get("sheet", "Sheet1"),
            "delay_min_seconds": 15,
            "delay_max_seconds": 30,
            "invalid_sheet_name": model.get("invalid_sheet", "邮箱失效"),
            "state_label": f"{model_key}_first_touch",
            "allow_create_tracking_column": False,
        },
        "workbook": {"path": str(model_workbook_path(config, model_key)).replace("\\", "/")},
        "sender": {
            "profiles_path": str(package / "sender_profiles.local.json").replace("\\", "/"),
            "profile": sender_profile or config.get("sender_profile", "qq_sender"),
        },
        "tracking": {
            "first_time_header": model.get("first_time_header", "一次跟进时间"),
            "first_reply_header": "一次跟进回复",
        },
        "field_aliases": {
            "company": [model.get("company_header", "")],
            "country": [model.get("country_header", "")],
            "email": [model.get("email_header", "")],
            "website": [model.get("website_header", "")],
            "first_time": [model.get("first_time_header", "")],
        },
        "templates": {
            "shell_path": str(package / "templates" / "email_shell.html").replace("\\", "/"),
            "signature_path": str(package / "templates" / "company_signature.html").replace("\\", "/"),
            "variables_path": str(package / "templates" / "variables.json").replace("\\", "/"),
        },
        "assets": {
            "logo_path": str(logo_file).replace("\\", "/"),
            "first_images": [str(path).replace("\\", "/") for path in image_paths],
        },
        "message": {
            "vehicle_model": model.get("message", {}).get("vehicle_model", model["label"]),
            "product_label": model.get("message", {}).get("product_label", model["label"] + " accessories"),
            "product_items": model.get("message", {}).get("product_items", ["accessories"]),
            "price_hint": "",
            "company_line_enabled": False,
        },
    }
    variables = read_json(package / "templates" / "variables.json", {})
    logo_path = config.get("logo_path") or variables.get("logoPath") or variables.get("logo_path") or ""
    if logo_path:
        raw_logo = Path(normalize(logo_path))
        logo_candidates = [raw_logo] if raw_logo.is_absolute() else [ROOT / raw_logo, package / "templates" / raw_logo]
        for candidate in logo_candidates:
            if candidate.exists():
                cfg["assets"]["logo_path"] = str(candidate).replace("\\", "/")
                break
    cfg_path = GENERATED_CONFIG_DIR / f"{model_key}.json"
    write_json(cfg_path, cfg)
    return cfg_path


def ensure_fallback_logo() -> Path:
    ASSET_DIR.mkdir(parents=True, exist_ok=True)
    logo = ASSET_DIR / "fallback-logo.png"
    if not logo.exists():
        logo.write_bytes(
            base64.b64decode(
                "iVBORw0KGgoAAAANSUhEUgAAAEAAAABACAYAAACqaXHeAAAByElEQVR4nO2aQU7DMBBF/9qBKoS4Qq6QK+QK3AEXoAogIs4Ayx5mpjHEm6YdS5P4nEiyHmxJYsdzHccxAAAAAAAAAAB8bO3s7m3sQH6GIo4jEwB7d3f3V7fb7X1x0f0OmW2sFcBv2w3AM4DnPiV5PxZlWQO8xXEcR3Z2dt4A3B3gk8lk/wDgbkKSJJvNZr8C8NwABsC+7+9vT0/PewH4C7C+Xq/fB+CuAK/X6/39/T0B3BXg6XS6v7+/p4C7AhydTnd3d/cEcEeA3+93e3t7O4A7AhwcHNze3t4G4A7A5XK5sbGxA7gjwMHBwcHBwQ7gjgD7+/s7OzsbwB0Bbm5uLi4uBvBAgMfHx0dHRwN4IsD29vbW1tYG8FCA1dXV9fX1AfAQgNPp9P7+/gDwUIDV1dUVFRUB8BCA9vb2+vr6APBQgKmpqampqQDwUICqqqqgoKAA8FCAjY2Njo6OAPBQgNfX1+vr6wDwUICrq6u3t7cA8FCAiYkJf39/AHhZgH6/39vb24C8K8D7/e7u7m4A8E6A8/nc3d0dAN4J8PLy8u7u7gDwPoC7u7tHR0cDeCZAZ2dnY2NjAHgmQHNzc2NjYwN4JkBPT0+Li4sD8AwAAAAAAAAAALyODxvNFj+aw+p5AAAAAElFTkSuQmCC"
            )
        )
    return logo


def send_for_model(model_key: str, limit: int, sender_profile: str | None = None) -> dict[str, Any]:
    config = load_config()
    model = config["models"][model_key]
    emails = unsent_emails_for_model(config, model_key, limit)
    if len(emails) < limit:
        raise RuntimeError(f"{model['label']} 未发客户只有 {len(emails)} 个，不能发送 {limit} 封")
    append_log(f"开始发送: {model['label']} limit={limit} targets={len(emails)}")
    results = []
    sent = 0
    failed = 0
    for email_value in emails:
        result = send_with_fallback_for_target(model_key, email_value, sender_profile, "手动发送")
        results.append({"email": email_value, **result})
        if result.get("ok"):
            sent += 1
        else:
            failed += 1
    append_log(f"完成发送: {model['label']} sent={sent} failed={failed}")
    return {
        "model": model_key,
        "label": model["label"],
        "limit": limit,
        "sender": sender_profile or config.get("sender_profile"),
        "sent": sent,
        "failed": failed,
        "results": results,
    }


def send_target_email_for_model(model_key: str, target_email: str, sender_profile: str | None = None, log_label: str = "退信重发") -> dict[str, Any]:
    config = load_config()
    model = config["models"][model_key]
    package = project_path(config["outreach_package"])
    retry_profile = sender_profile or normalize(config.get("retry_sender_profile")) or config.get("sender_profile", "qq_sender")
    cfg_path = ensure_outreach_config(config, model_key, retry_profile)
    python_exe = python_executable(config)
    script = package / "scripts" / "send_portable_campaign.py"
    cmd = [
        python_exe,
        str(script),
        "--config",
        str(cfg_path),
        "--limit",
        "1",
        "--target-emails",
        target_email,
        "--force-targets",
    ]
    append_log(f"{log_label}: {model['label']} {target_email} sender={retry_profile}")
    proc = subprocess.run(cmd, cwd=str(package / "scripts"), capture_output=True, text=True, encoding="utf-8", errors="replace")
    if proc.stdout:
        append_log(proc.stdout.strip()[-3000:])
    if proc.stderr:
        append_log(f"{log_label}错误输出: " + proc.stderr.strip()[-3000:])
    if proc.returncode != 0:
        raise RuntimeError(f"{log_label}失败，退出码 {proc.returncode}: {proc.stderr[-1000:]}")
    return {"model": model_key, "label": model["label"], "email": target_email, "returncode": proc.returncode}


def copy_cell_value_and_style(source: Any, target: Any) -> None:
    target.value = source.value
    if source.has_style:
        target._style = copy(source._style)
    target.number_format = source.number_format


def ensure_label_column(sheet: Any, headers: list[Any], names: list[str]) -> int:
    label_col = header_index(headers, names)
    if label_col:
        return label_col
    label_col = sheet.max_column + 1
    sheet.cell(1, label_col).value = names[0]
    return label_col


def copy_row_to_destination(
    workbook: Any,
    source_sheet: Any,
    row_number: int,
    destination_name: str,
    label: str,
    model: dict[str, Any],
    label_headers: list[str],
    note: str = "",
) -> None:
    if destination_name in workbook.sheetnames:
        destination = workbook[destination_name]
    else:
        destination = workbook.create_sheet(destination_name)
        for col in range(1, source_sheet.max_column + 1):
            copy_cell_value_and_style(source_sheet.cell(1, col), destination.cell(1, col))
        destination.cell(1, source_sheet.max_column + 1).value = label_headers[0]
        for letter, dim in source_sheet.column_dimensions.items():
            destination.column_dimensions[letter].width = dim.width

    dest_row = destination.max_row + 1
    source_headers = headers_for_sheet(source_sheet)
    dest_headers = headers_for_sheet(destination)
    aliases = field_aliases_for_model(model)
    source_map = {field: header_index(source_headers, names) for field, names in aliases.items()}
    dest_map = {field: header_index(dest_headers, names) for field, names in aliases.items()}
    mapped_fields = [field for field in aliases if source_map.get(field) and dest_map.get(field)]

    if mapped_fields:
        for field in mapped_fields:
            copy_cell_value_and_style(source_sheet.cell(row_number, source_map[field]), destination.cell(dest_row, dest_map[field]))
    else:
        for col in range(1, source_sheet.max_column + 1):
            copy_cell_value_and_style(source_sheet.cell(row_number, col), destination.cell(dest_row, col))

    dest_headers = headers_for_sheet(destination)
    label_col = ensure_label_column(destination, dest_headers, label_headers + ["标签", "备注", "一次跟进回复", "回复"])
    destination.cell(dest_row, label_col).value = label

    dest_headers = headers_for_sheet(destination)
    source_sheet_col = header_index(dest_headers, ["来源Sheet", "source sheet", "source_sheet"])
    if source_sheet_col:
        destination.cell(dest_row, source_sheet_col).value = source_sheet.title
    source_row_col = header_index(dest_headers, ["原行号", "source row", "source_row"])
    if source_row_col:
        destination.cell(dest_row, source_row_col).value = row_number
    note_col = header_index(dest_headers, ["备注", "note", "notes"])
    if note and note_col:
        destination.cell(dest_row, note_col).value = note


def copy_row_to_invalid(workbook: Any, sheet: Any, row_number: int, invalid_name: str, label: str, model: dict[str, Any]) -> None:
    copy_row_to_destination(workbook, sheet, row_number, invalid_name, label, model, ["标签", "失效原因", "退信原因"])


def move_invalid_email(config: dict[str, Any], model_key: str, email_value: str, label: str = "退信") -> bool:
    model = config["models"][model_key]
    path, workbook, sheet, _headers, mapping = open_model_sheet(config, model_key, data_only=False)
    moved = False
    try:
        email_col = mapping.get("email")
        if not email_col:
            return False
        destination_name = resolve_destination_sheet_name(workbook, sheet, model.get("invalid_sheet", "邮箱失效"), "邮箱失效", "invalid")
        for row_number in range(sheet.max_row, 1, -1):
            value = normalize(sheet.cell(row_number, email_col).value).lower()
            if value == email_value.lower():
                copy_row_to_invalid(workbook, sheet, row_number, destination_name, label, model)
                sheet.delete_rows(row_number, 1)
                moved = True
        if moved:
            workbook.save(path)
            append_log(f"已移入失效邮箱: {model['label']} {email_value}")
        return moved
    finally:
        workbook.close()


def mark_final_invalid_email(config: dict[str, Any], model_key: str, email_value: str, sender_profile: str, reason: str, label: str) -> bool:
    record_today_invalid(model_key, email_value, sender_profile, reason)
    try:
        moved = move_invalid_email(config, model_key, email_value, label)
    except Exception as exc:
        append_log(f"移入邮箱失效失败: {email_value} :: {exc}")
        return False
    append_log(f"最终失效处理: {email_value} moved={moved}")
    return moved


def move_followup_email(config: dict[str, Any], model_key: str, email_value: str, label: str = "已回复客户", note: str = "") -> bool:
    model = config["models"][model_key]
    path, workbook, sheet, _headers, mapping = open_model_sheet(config, model_key, data_only=False)
    moved = False
    try:
        email_col = mapping.get("email")
        if not email_col:
            return False
        destination_name = resolve_destination_sheet_name(workbook, sheet, model.get("followup_sheet", ""), "跟进", "followup")
        for row_number in range(sheet.max_row, 1, -1):
            value = normalize(sheet.cell(row_number, email_col).value).lower()
            if value == email_value.lower():
                copy_row_to_destination(
                    workbook,
                    sheet,
                    row_number,
                    destination_name,
                    label,
                    model,
                    ["标签", "一次跟进回复", "回复", "备注"],
                    note=note,
                )
                sheet.delete_rows(row_number, 1)
                moved = True
        if moved:
            workbook.save(path)
            append_log(f"已移入跟进sheet: {model['label']} {email_value} -> {destination_name}")
        return moved
    finally:
        workbook.close()


def find_model_for_email(config: dict[str, Any], email_value: str) -> str | None:
    target = email_value.lower()
    for key in config.get("models", {}):
        if any(row.get("email", "").lower() == target for row in read_model_rows(config, key)):
            return key
    return None


def find_model_email_for_candidates(config: dict[str, Any], emails: list[str]) -> tuple[str | None, str | None]:
    seen = []
    for email_value in emails:
        normalized = normalize(email_value).lower()
        if normalized and normalized not in seen:
            seen.append(normalized)
    for email_value in seen:
        model_key = find_model_for_email(config, email_value)
        if model_key:
            return model_key, email_value
    return None, None


def find_unsent_alternative(config: dict[str, Any], model_key: str, exclude_email: str) -> str | None:
    for row in read_model_rows(config, model_key):
        if row.get("email", "").lower() == exclude_email.lower():
            continue
        if not row.get("first_time"):
            return row.get("email")
    return None


def unsent_emails_for_model(config: dict[str, Any], model_key: str, limit: int) -> list[str]:
    emails = []
    for row in read_model_rows(config, model_key):
        if row.get("first_time"):
            continue
        email_value = normalize(row.get("email")).lower()
        if email_value and email_value not in emails:
            emails.append(email_value)
        if len(emails) >= limit:
            break
    return emails


def retry_after_bounce(config: dict[str, Any], model_key: str, bounced_email: str) -> dict[str, Any]:
    state = read_json(BOUNCE_STATE_PATH, {})
    key = bounced_email.lower()
    if state.get(key, {}).get("retried"):
        state[key] = {"retried": True, "failed_final": True, "updated_at": now_iso()}
        write_json(BOUNCE_STATE_PATH, state)
        append_log(f"退信二次失败忽略: {bounced_email}")
        return {"action": "ignored_retried_bounce"}
    retry_sender = normalize(config.get("retry_sender_profile")) or normalize(config.get("sender_profile"))
    try:
        send_result = send_target_email_for_model(model_key, bounced_email, retry_sender)
    except Exception as exc:
        state[key] = {"retried": True, "failed_final": True, "updated_at": now_iso(), "retry_error": str(exc)[:500]}
        write_json(BOUNCE_STATE_PATH, state)
        moved = mark_final_invalid_email(config, model_key, bounced_email, retry_sender, str(exc), "退信备用发送失败")
        append_log(f"退信备用发送失败，已计入今日失效: {bounced_email} moved={moved}")
        return {"action": "retry_send_failed", "error": str(exc), "moved_to_invalid": moved}
    state[key] = {"retried": True, "failed_final": False, "updated_at": now_iso()}
    write_json(BOUNCE_STATE_PATH, state)
    return {"action": "retried_once", "send_result": send_result}


def decode_payload(payload: bytes, charset: str | None) -> str:
    encoding = (charset or "utf-8").lower()
    if encoding in {"unknown-8bit", "x-unknown", "unknown"}:
        encoding = "utf-8"
    try:
        return payload.decode(encoding, errors="replace")
    except LookupError:
        return payload.decode("utf-8", errors="replace")


def message_text(msg: email.message.Message) -> str:
    parts = []
    if msg.is_multipart():
        for part in msg.walk():
            ctype = part.get_content_type()
            if ctype not in {"text/plain", "text/html"}:
                continue
            payload = part.get_payload(decode=True)
            if payload:
                parts.append(decode_payload(payload, part.get_content_charset()))
    else:
        payload = msg.get_payload(decode=True)
        if payload:
            parts.append(decode_payload(payload, msg.get_content_charset()))
    return "\n".join(parts)


def clean_message_text(text: str) -> str:
    cleaned = re.sub(r"(?is)<(script|style).*?</\1>", " ", text)
    cleaned = re.sub(r"(?is)<br\s*/?>", "\n", cleaned)
    cleaned = re.sub(r"(?is)</p\s*>", "\n", cleaned)
    cleaned = re.sub(r"(?is)<[^>]+>", " ", cleaned)
    cleaned = unescape(cleaned)
    cleaned = re.sub(r"\r\n?", "\n", cleaned)
    cleaned = re.sub(r"[ \t]+", " ", cleaned)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    return cleaned.strip()


def sender_email(from_addr: str) -> str:
    matches = EMAIL_RE.findall(from_addr or "")
    return matches[0].lower() if matches else ""


def own_sender_emails(config: dict[str, Any]) -> set[str]:
    emails: set[str] = set()
    for account in config.get("mail_accounts", []):
        resolved = resolve_mail_account(config, account)
        if resolved.get("imap_user"):
            emails.add(resolved["imap_user"].lower())
    for profile in sender_profiles(config).values():
        user = normalize(profile.get("user")).lower()
        if user:
            emails.add(user)
    return emails


def leading_message_text(text: str) -> str:
    head = clean_message_text(text)
    delimiters = (
        "\nOn ",
        "\nFrom:",
        "\nDe :",
        "\n发件人：",
        "\n-----Original Message-----",
        "\n________________________________",
    )
    for delimiter in delimiters:
        if delimiter in head:
            head = head.split(delimiter, 1)[0]
    return head.strip()


def is_auto_or_noise(subject: str, from_addr: str, text: str) -> bool:
    sender = sender_email(from_addr)
    local = sender.split("@", 1)[0] if sender else ""
    merged = f"{subject}\n{from_addr}\n{text}".lower()
    if local in {"noreply", "no-reply", "donotreply", "do-not-reply", "mailer-daemon", "postmaster"}:
        return True
    return any(term in merged for term in AUTO_REPLY_TERMS)


def is_interested_message(config: dict[str, Any], subject: str, from_addr: str, text: str) -> bool:
    lead = leading_message_text(text)
    if not lead:
        return False
    if is_auto_or_noise(subject, from_addr, lead):
        return False
    merged = f"{subject}\n{lead}".lower()
    return any(re.search(pattern, merged, re.I) for pattern in STRONG_INTEREST_PATTERNS)


def intervention_id(item: dict[str, Any]) -> str:
    raw = "|".join(
        [
            normalize(item.get("from")),
            normalize(item.get("subject")),
            normalize(item.get("email")),
            normalize(item.get("time")),
        ]
    )
    return hashlib.sha1(raw.encode("utf-8", errors="replace")).hexdigest()[:16]


def normalize_intervention_item(item: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(item)
    normalized.setdefault("id", intervention_id(normalized))
    normalized.setdefault("read_at", "")
    if normalized.get("body") and not normalized.get("snippet"):
        normalized["snippet"] = clean_message_text(normalized["body"])[:240]
    return normalized


def read_interventions() -> list[dict[str, Any]]:
    raw = read_json(INTERVENTION_PATH, [])
    if isinstance(raw, dict):
        if isinstance(raw.get("value"), list):
            raw = raw["value"]
        elif raw:
            raw = [raw]
        else:
            raw = []
    if not isinstance(raw, list):
        raw = []
    return [normalize_intervention_item(item) for item in raw if isinstance(item, dict)]


def add_intervention(item: dict[str, Any]) -> None:
    item = normalize_intervention_item(item)
    items = read_interventions()
    signature = (item.get("from", ""), item.get("subject", ""), item.get("email", ""))
    for existing in items:
        if (existing.get("from", ""), existing.get("subject", ""), existing.get("email", "")) == signature:
            return
    items.insert(0, item)
    write_json(INTERVENTION_PATH, items[:300])


def update_intervention(item_id: str, values: dict[str, Any]) -> bool:
    items = read_interventions()
    changed = False
    for item in items:
        if item.get("id") == item_id:
            item.update(values)
            changed = True
            break
    if changed:
        write_json(INTERVENTION_PATH, items)
    return changed


def clear_active_interventions() -> int:
    items = read_interventions()
    cleared = 0
    for item in items:
        if item.get("status") != "已处理":
            item["status"] = "已处理"
            item["handled_at"] = now_iso()
            item["cleared"] = True
            cleared += 1
    write_json(INTERVENTION_PATH, items)
    return cleared


def mail_message_id(account_key: str, folder: str, msg_id: str, subject: str, from_addr: str, uid: str = "") -> str:
    raw = "|".join([account_key, folder, normalize(uid) or msg_id, normalize(subject), normalize(from_addr)])
    return hashlib.sha1(raw.encode("utf-8", errors="replace")).hexdigest()[:16]


def record_mail_message(account_key: str, folder: str, msg_id: str, subject: str, from_addr: str, text: str, uid: str = "") -> None:
    messages = read_json(MAIL_MESSAGES_PATH, {})
    if not isinstance(messages, dict):
        messages = {}
    body = leading_message_text(text) or clean_message_text(text)
    item = {
        "id": mail_message_id(account_key, folder, msg_id, subject, from_addr, uid),
        "time": now_iso(),
        "account_key": account_key,
        "folder": folder,
        "imap_uid": normalize(uid),
        "imap_id": msg_id,
        "from": from_addr,
        "subject": subject or "无主题",
        "snippet": re.sub(r"\s+", " ", body)[:180],
        "body": body[:12000],
    }
    account_messages = messages.get(account_key, [])
    if not isinstance(account_messages, list):
        account_messages = []
    for existing in account_messages:
        if existing.get("id") == item["id"] and existing.get("read_at"):
            item["read_at"] = existing["read_at"]
            break
    account_messages = [existing for existing in account_messages if existing.get("id") != item["id"]]
    account_messages.insert(0, item)
    messages[account_key] = account_messages[:20]
    write_json(MAIL_MESSAGES_PATH, messages)


def mark_imap_message_seen(config: dict[str, Any], account_key: str, message: dict[str, Any]) -> bool:
    account = next((item for item in config.get("mail_accounts", []) if normalize(item.get("key")) == account_key), None)
    if not account:
        return False
    resolved = resolve_mail_account(config, account)
    folder = normalize(message.get("folder")) or "INBOX"
    uid = normalize(message.get("imap_uid"))
    seq_id = normalize(message.get("imap_id"))
    if not (resolved["imap_host"] and resolved["imap_user"] and resolved["imap_password"] and folder and (uid or seq_id)):
        return False
    with imaplib.IMAP4_SSL(resolved["imap_host"], resolved["imap_port"], timeout=45) as client:
        client.login(resolved["imap_user"], resolved["imap_password"])
        select_status, _ = client.select(folder, readonly=False)
        if select_status != "OK":
            return False
        if uid:
            status, _ = client.uid("STORE", uid, "+FLAGS", r"(\Seen)")
        else:
            status, _ = client.store(seq_id, "+FLAGS", r"(\Seen)")
        return status == "OK"


def mark_mail_message_read(account_key: str, message_id: str) -> dict[str, Any]:
    messages = read_json(MAIL_MESSAGES_PATH, {})
    if not isinstance(messages, dict):
        return {"updated": False, "imap_seen": False}
    keys = [account_key] if account_key else list(messages.keys())
    for key in keys:
        account_messages = messages.get(key, [])
        if not isinstance(account_messages, list):
            continue
        for item in account_messages:
            if item.get("id") == message_id:
                imap_seen = False
                imap_error = ""
                try:
                    imap_seen = mark_imap_message_seen(load_config(), key, item)
                except Exception as exc:
                    imap_error = str(exc)
                    append_log(f"同步邮件已读失败: {key} {message_id}: {exc}")
                item["read_at"] = now_iso()
                write_json(MAIL_MESSAGES_PATH, messages)
                return {"updated": True, "imap_seen": imap_seen, "imap_error": imap_error}
    return {"updated": False, "imap_seen": False}


def remove_mail_message(account_key: str, message_id: str) -> dict[str, Any]:
    messages = read_json(MAIL_MESSAGES_PATH, {})
    if not isinstance(messages, dict):
        return {"removed": False, "imap_seen": False}
    keys = [account_key] if account_key else list(messages.keys())
    for key in keys:
        account_messages = messages.get(key, [])
        if not isinstance(account_messages, list):
            continue
        target = next((item for item in account_messages if item.get("id") == message_id), None)
        imap_seen = False
        imap_error = ""
        if target:
            try:
                imap_seen = mark_imap_message_seen(load_config(), key, target)
            except Exception as exc:
                imap_error = str(exc)
                append_log(f"同步邮件已读失败: {key} {message_id}: {exc}")
        filtered = [item for item in account_messages if item.get("id") != message_id]
        if len(filtered) != len(account_messages):
            messages[key] = filtered
            write_json(MAIL_MESSAGES_PATH, messages)
            return {"removed": True, "imap_seen": imap_seen, "imap_error": imap_error}
    return {"removed": False, "imap_seen": False}


def mail_match_text(value: Any) -> str:
    return re.sub(r"\s+", " ", normalize(value)).strip().lower()


def remove_matching_mail_messages(intervention: dict[str, Any]) -> list[dict[str, str]]:
    messages = read_json(MAIL_MESSAGES_PATH, {})
    if not isinstance(messages, dict):
        return []
    target_from = mail_match_text(intervention.get("from"))
    target_subject = mail_match_text(intervention.get("subject"))
    target_email = mail_match_text(intervention.get("email"))
    target_body = mail_match_text(intervention.get("snippet") or intervention.get("body"))
    removed: list[dict[str, str]] = []
    changed = False
    for account_key, account_messages in list(messages.items()):
        if not isinstance(account_messages, list):
            continue
        kept = []
        for item in account_messages:
            item_from = mail_match_text(item.get("from"))
            item_subject = mail_match_text(item.get("subject"))
            item_body = mail_match_text(item.get("snippet") or item.get("body"))
            from_matches = bool(target_from and item_from == target_from) or bool(target_email and target_email in item_from)
            subject_matches = bool(target_subject and item_subject == target_subject)
            body_matches = bool(
                target_body
                and item_body
                and (target_body[:120] in item_body or item_body[:120] in target_body)
            )
            if subject_matches and (from_matches or body_matches):
                removed.append({"account_key": account_key, "id": normalize(item.get("id"))})
                changed = True
                continue
            kept.append(item)
        messages[account_key] = kept
    if changed:
        write_json(MAIL_MESSAGES_PATH, messages)
    return removed


def poll_mail_once() -> dict[str, Any]:
    config = load_config()
    mail_cfg = config.get("mail_monitor", {})
    if not mail_cfg.get("imap_host") or not mail_cfg.get("imap_user") or not mail_cfg.get("imap_password"):
        return {"ok": False, "message": "IMAP 未配置"}
    processed = read_json(RUNTIME_DIR / "mail_processed.json", {})
    since = (datetime.now() - timedelta(days=int(mail_cfg.get("lookback_days", 14)))).strftime("%d-%b-%Y")
    with imaplib.IMAP4_SSL(mail_cfg["imap_host"], int(mail_cfg.get("imap_port", 993))) as client:
        client.login(mail_cfg["imap_user"], mail_cfg["imap_password"])
        client.select(mail_cfg.get("mailbox", "INBOX"))
        status, data = client.search(None, f'(SINCE "{since}")')
        if status != "OK":
            return {"ok": False, "message": "IMAP search failed"}
        ids = data[0].split()
        checked = 0
        bounces = 0
        interested = 0
        for msg_id in ids[-200:]:
            msg_key = msg_id.decode("ascii", errors="ignore")
            if processed.get(msg_key):
                continue
            status, msg_data = client.fetch(msg_id, "(RFC822)")
            if status != "OK" or not msg_data or not isinstance(msg_data[0], tuple):
                continue
            raw = msg_data[0][1]
            msg = email.message_from_bytes(raw)
            subject = decode_mime_header(msg.get("Subject"))
            from_addr = decode_mime_header(msg.get("From"))
            text = message_text(msg)
            lower = f"{subject}\n{from_addr}\n{text}".lower()
            checked += 1
            if any(token.lower() in lower for token in config.get("bounce_keywords", [])):
                emails = [value.lower() for value in EMAIL_RE.findall(text) if is_valid_email(value)]
                for bounced in dict.fromkeys(emails):
                    model_key = find_model_for_email(config, bounced)
                    if model_key:
                        retry_after_bounce(config, model_key, bounced)
                        bounces += 1
                        break
            elif any(token.lower() in lower for token in config.get("interest_keywords", [])):
                emails = [value.lower() for value in EMAIL_RE.findall(f"{from_addr}\n{text}") if is_valid_email(value)]
                item = {
                    "time": now_iso(),
                    "from": from_addr,
                    "subject": subject,
                    "email": emails[0] if emails else "",
                    "snippet": re.sub(r"\s+", " ", text)[:500],
                    "status": "待处理",
                }
                add_intervention(item)
                interested += 1
            processed[msg_key] = now_iso()
        write_json(RUNTIME_DIR / "mail_processed.json", processed)
        return {"ok": True, "checked": checked, "bounces": bounces, "interested": interested}


def mail_monitor_loop() -> None:
    append_log("邮件监控已启动")
    while not MAIL_MONITOR_STOP.is_set():
        try:
            result = poll_mail_once()
            append_log("邮件监控: " + json.dumps(result, ensure_ascii=False))
        except Exception as exc:
            append_log(f"邮件监控错误: {exc}")
        interval = int(load_config().get("mail_monitor", {}).get("interval_seconds", 300))
        MAIL_MONITOR_STOP.wait(max(30, interval))
    append_log("邮件监控已停止")


def start_mail_monitor() -> dict[str, Any]:
    global MAIL_MONITOR_THREAD
    mail_cfg = load_config().get("mail_monitor", {})
    if not (mail_cfg.get("imap_host") and mail_cfg.get("imap_user") and mail_cfg.get("imap_password")):
        return {"running": False, "message": "IMAP 未配置"}
    if MAIL_MONITOR_THREAD is not None and MAIL_MONITOR_THREAD.is_alive():
        return {"running": True, "message": "already running"}
    MAIL_MONITOR_STOP.clear()
    MAIL_MONITOR_THREAD = threading.Thread(target=mail_monitor_loop, name="mail-monitor", daemon=True)
    MAIL_MONITOR_THREAD.start()
    return {"running": True}


def stop_mail_monitor() -> dict[str, Any]:
    MAIL_MONITOR_STOP.set()
    return {"running": False}


def process_mail_message(config: dict[str, Any], raw: bytes) -> tuple[int, int]:
    msg = email.message_from_bytes(raw)
    subject = decode_mime_header(msg.get("Subject"))
    from_addr = decode_mime_header(msg.get("From"))
    text = message_text(msg)
    lower = f"{subject}\n{from_addr}\n{text}".lower()
    if any(token.lower() in lower for token in config.get("bounce_keywords", [])):
        if should_ignore_bounce(config, msg):
            return 0, 0
        emails = [value.lower() for value in EMAIL_RE.findall(text) if is_valid_email(value)]
        for bounced in dict.fromkeys(emails):
            model_key = find_model_for_email(config, bounced)
            if model_key:
                retry_after_bounce(config, model_key, bounced)
                return 1, 0
        return 0, 0
    if is_interested_message(config, subject, from_addr, text):
        emails = [value.lower() for value in EMAIL_RE.findall(f"{from_addr}\n{text}") if is_valid_email(value)]
        model_key, customer_email = find_model_email_for_candidates(config, emails)
        moved_to_followup = False
        if model_key and customer_email:
            moved_to_followup = move_followup_email(config, model_key, customer_email, "已回复客户", "检测到客户意向回复")
        body = leading_message_text(text)
        add_intervention(
            {
                "time": now_iso(),
                "from": from_addr,
                "subject": subject,
                "email": customer_email or (emails[0] if emails else ""),
                "model": model_key or "",
                "moved_to_followup": moved_to_followup,
                "snippet": re.sub(r"\s+", " ", body)[:240],
                "body": body[:12000],
                "status": "待处理",
            }
        )
        return 0, 1
    return 0, 0


def poll_mail_account(config: dict[str, Any], account: dict[str, Any], processed: dict[str, str]) -> dict[str, Any]:
    resolved = resolve_mail_account(config, account)
    status = {
        "key": resolved["key"],
        "label": resolved["label"],
        "user": resolved["imap_user"],
        "status": "未配置",
        "last_checked": now_iso(),
        "checked": 0,
        "bounces": 0,
        "interested": 0,
        "folders": [],
        "error": "",
    }
    if not (resolved["imap_host"] and resolved["imap_user"] and resolved["imap_password"]) or resolved["imap_password"] == "replace-me":
        status["error"] = "IMAP账号未配置"
        return status
    since = (datetime.now() - timedelta(days=int(config.get("mail_monitor", {}).get("lookback_days", 14)))).strftime("%d-%b-%Y")
    try:
        with imaplib.IMAP4_SSL(resolved["imap_host"], resolved["imap_port"], timeout=45) as client:
            client.login(resolved["imap_user"], resolved["imap_password"])
            for folder in resolved["folders"]:
                try:
                    select_status, _ = client.select(folder, readonly=True)
                except Exception:
                    continue
                if select_status != "OK":
                    continue
                status["folders"].append(folder)
                search_status, data = client.uid("SEARCH", None, f'(UNSEEN SINCE "{since}")')
                if search_status != "OK":
                    continue
                for uid_value in data[0].split()[-200:]:
                    uid_text = uid_value.decode("ascii", errors="ignore")
                    msg_key = f"{resolved['key']}:{folder}:UID:{uid_text}"
                    if processed.get(msg_key):
                        continue
                    fetch_status, msg_data = client.uid("FETCH", uid_value, "(RFC822)")
                    if fetch_status != "OK" or not msg_data or not isinstance(msg_data[0], tuple):
                        continue
                    raw = msg_data[0][1]
                    msg = email.message_from_bytes(raw)
                    subject = decode_mime_header(msg.get("Subject"))
                    from_addr = decode_mime_header(msg.get("From"))
                    text = message_text(msg)
                    record_mail_message(resolved["key"], folder, uid_text, subject, from_addr, text, uid=uid_text)
                    bounces, interested = process_mail_message(config, raw)
                    status["checked"] += 1
                    status["bounces"] += bounces
                    status["interested"] += interested
                    processed[msg_key] = now_iso()
        status["status"] = "正常"
    except Exception as exc:
        status["status"] = "异常"
        status["error"] = str(exc)[:300]
    return status


def poll_mail_once() -> dict[str, Any]:
    config = load_config()
    processed = read_json(RUNTIME_DIR / "mail_processed.json", {})
    statuses = {}
    total = {"checked": 0, "bounces": 0, "interested": 0}
    for account in config.get("mail_accounts", []):
        result = poll_mail_account(config, account, processed)
        statuses[result["key"]] = result
        total["checked"] += int(result.get("checked", 0))
        total["bounces"] += int(result.get("bounces", 0))
        total["interested"] += int(result.get("interested", 0))
    write_json(RUNTIME_DIR / "mail_processed.json", processed)
    write_json(MAIL_STATUS_PATH, statuses)
    return {"ok": True, "accounts": list(statuses.values()), **total}


def start_mail_monitor() -> dict[str, Any]:
    global MAIL_MONITOR_THREAD
    if MAIL_MONITOR_THREAD is not None and MAIL_MONITOR_THREAD.is_alive():
        return {"running": True, "message": "already running"}
    MAIL_MONITOR_STOP.clear()
    MAIL_MONITOR_THREAD = threading.Thread(target=mail_monitor_loop, name="mail-monitor", daemon=True)
    MAIL_MONITOR_THREAD.start()
    return {"running": True}


def run_task(kind: str, fn: Any, *args: Any) -> dict[str, Any]:
    global CURRENT_TASK
    with TASK_LOCK:
        if CURRENT_TASK and CURRENT_TASK.get("status") == "running":
            raise RuntimeError("已有任务正在运行")
        CURRENT_TASK = {"kind": kind, "status": "running", "started_at": now_iso(), "progress": "", "result": None}

    def worker() -> None:
        global CURRENT_TASK
        try:
            result = fn(*args)
            CURRENT_TASK = {**(CURRENT_TASK or {}), "status": "done", "finished_at": now_iso(), "result": result}
            append_log(f"任务完成: {kind}")
        except Exception as exc:
            CURRENT_TASK = {**(CURRENT_TASK or {}), "status": "failed", "finished_at": now_iso(), "error": str(exc)}
            append_log(f"任务失败: {kind}: {exc}")

    threading.Thread(target=worker, name=f"task-{kind}", daemon=True).start()
    return CURRENT_TASK


class ApiHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args: Any, **kwargs: Any) -> None:
        super().__init__(*args, directory=str(STATIC_DIR), **kwargs)

    def log_message(self, format: str, *args: Any) -> None:
        append_log("HTTP " + format % args)

    def read_body(self) -> dict[str, Any]:
        length = int(self.headers.get("Content-Length", "0") or "0")
        if length <= 0:
            return {}
        return json.loads(self.rfile.read(length).decode("utf-8"))

    def send_json(self, data: Any, status: int = 200) -> None:
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def end_headers(self) -> None:
        self.send_header("Cache-Control", "no-store, max-age=0")
        super().end_headers()

    def do_GET(self) -> None:
        if self.path == "/api/status":
            self.send_json(dashboard_status())
            return
        if self.path == "/api/config":
            cfg = load_config()
            redacted = dict(cfg)
            if redacted.get("anysearch_api_key"):
                redacted["anysearch_api_key"] = "***"
            mail_cfg = dict(redacted.get("mail_monitor", {}))
            if mail_cfg.get("imap_password"):
                mail_cfg["imap_password"] = "***"
            redacted["mail_monitor"] = mail_cfg
            redacted["mail_accounts"] = [
                {**account, "imap_password": "***" if account.get("imap_password") else ""}
                for account in redacted.get("mail_accounts", [])
            ]
            self.send_json(redacted)
            return
        super().do_GET()

    def do_POST(self) -> None:
        try:
            if self.path == "/api/daily-collect":
                body = self.read_body()
                limit = max(1, min(50, int(body.get("limit", 10))))
                self.send_json(run_task("daily_collect", daily_collect, limit))
                return
            if self.path == "/api/collect-preview":
                body = self.read_body()
                model_key = body.get("model")
                config = load_config()
                if model_key not in config.get("models", {}):
                    raise RuntimeError("未知车型")
                limit = max(1, min(20, int(body.get("limit", 10))))
                query_limit = max(1, min(20, int(body.get("query_limit", 8))))
                self.send_json(preview_collect_for_model(config, model_key, limit, query_limit))
                return
            if self.path == "/api/send":
                body = self.read_body()
                model_key = body.get("model")
                limit = max(0, min(500, int(body.get("limit", 0))))
                sender_profile = normalize(body.get("sender"))
                if model_key not in load_config().get("models", {}):
                    raise RuntimeError("未知车型")
                if limit <= 0:
                    raise RuntimeError("发送数量必须大于 0")
                self.send_json(run_task(f"send_{model_key}", send_for_model, model_key, limit, sender_profile))
                return
            if self.path == "/api/reset-model":
                body = self.read_body()
                model_key = body.get("model")
                if model_key not in load_config().get("models", {}):
                    raise RuntimeError("未知车型")
                self.send_json(run_task(f"reset_{model_key}", reset_model_send_state, model_key))
                return
            if self.path == "/api/mail/start":
                self.send_json(start_mail_monitor())
                return
            if self.path == "/api/mail/stop":
                self.send_json(stop_mail_monitor())
                return
            if self.path == "/api/mail/poll":
                self.send_json(run_task("mail_poll_once", poll_mail_once))
                return
            if self.path == "/api/mail/read":
                body = self.read_body()
                message_id = normalize(body.get("id"))
                if not message_id:
                    raise RuntimeError("缺少邮件ID")
                result = mark_mail_message_read(normalize(body.get("account_key")), message_id)
                self.send_json({"ok": True, **result})
                return
            if self.path == "/api/mail/remove":
                body = self.read_body()
                message_id = normalize(body.get("id"))
                if not message_id:
                    raise RuntimeError("缺少邮件ID")
                result = remove_mail_message(normalize(body.get("account_key")), message_id)
                self.send_json({"ok": True, **result})
                return
            if self.path == "/api/intervention/close":
                body = self.read_body()
                item_id = normalize(body.get("id"))
                if item_id:
                    target = next((item for item in read_interventions() if item.get("id") == item_id), None)
                    updated = update_intervention(item_id, {"status": "已处理", "handled_at": now_iso()})
                    removed_mail_messages = remove_matching_mail_messages(target or {}) if updated else []
                    self.send_json({"ok": True, "updated": updated, "removed_mail_messages": removed_mail_messages})
                    return
                index = int(body.get("index", -1))
                items = read_interventions()
                removed_mail_messages = []
                if 0 <= index < len(items):
                    removed_mail_messages = remove_matching_mail_messages(items[index])
                    items[index]["status"] = "已处理"
                    items[index]["handled_at"] = now_iso()
                    write_json(INTERVENTION_PATH, items)
                self.send_json({"ok": True, "updated": 0 <= index < len(items), "removed_mail_messages": removed_mail_messages})
                return
            if self.path == "/api/intervention/read":
                body = self.read_body()
                item_id = normalize(body.get("id"))
                if not item_id:
                    raise RuntimeError("缺少邮件ID")
                updated = update_intervention(item_id, {"read_at": now_iso()})
                self.send_json({"ok": True, "updated": updated})
                return
            if self.path == "/api/intervention/clear":
                cleared = clear_active_interventions()
                self.send_json({"ok": True, "cleared": cleared})
                return
            self.send_json({"error": "not found"}, 404)
        except Exception as exc:
            self.send_json({"error": str(exc)}, 500)


def main() -> int:
    RUNTIME_DIR.mkdir(parents=True, exist_ok=True)
    STATIC_DIR.mkdir(parents=True, exist_ok=True)
    config = load_config()
    mail_cfg = config.get("mail_monitor", {})
    if mail_cfg.get("enabled_on_start"):
        start_mail_monitor()
    host = config.get("host", "127.0.0.1")
    port = int(config.get("port", 8765))
    append_log(f"客户增长控制台启动 http://{host}:{port}")
    server = ThreadingHTTPServer((host, port), ApiHandler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        stop_mail_monitor()
        server.server_close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
