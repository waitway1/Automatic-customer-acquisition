#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import smtplib
import time
from datetime import datetime
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from outreach_copy_engine import OutreachCustomer, analyze_site, build_lines, choose_products, normalize, subject_for


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", required=True)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--test-to", default="")
    parser.add_argument("--target-rows", default="")
    parser.add_argument("--target-emails", default="")
    parser.add_argument("--force-targets", action="store_true")
    return parser.parse_args()


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def save_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def resolve_path(base_dir: Path, value: str) -> Path:
    path = Path(value)
    return path if path.is_absolute() else (base_dir / path).resolve()


def compact(value: Any) -> str:
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", normalize(value).lower())


def map_headers(headers: list[Any], aliases: dict[str, list[str]]) -> dict[str, int]:
    mapped: dict[str, int] = {}
    normalized = [compact(item) for item in headers]
    for key, values in aliases.items():
        wanted = [compact(item) for item in values if normalize(item)]
        for index, header in enumerate(normalized):
            if any(alias == header for alias in wanted):
                mapped[key] = index + 1
                break
        if key in mapped:
            continue
        for index, header in enumerate(normalized):
            if any(alias and alias in header for alias in wanted):
                mapped[key] = index + 1
                break
    return mapped


def field_aliases(config: dict[str, Any]) -> dict[str, list[str]]:
    aliases = {key: list(values) for key, values in config.get("field_aliases", {}).items()}
    tracking = config.get("tracking", {})
    if tracking.get("first_time_header"):
        aliases.setdefault("first_time", []).append(tracking["first_time_header"])
    if tracking.get("first_reply_header"):
        aliases.setdefault("first_reply", []).append(tracking["first_reply_header"])
    return aliases


def cell_value(sheet: Any, row: int, header_map: dict[str, int], key: str) -> str:
    column = header_map.get(key)
    return normalize(sheet.cell(row, column).value) if column else ""


def parse_target_emails(value: str) -> set[str]:
    parts = re.split(r"[,;\s]+", normalize(value).lower())
    return {part for part in parts if "@" in part}


def parse_target_rows(value: str) -> set[int]:
    rows: set[int] = set()
    for part in re.split(r"[,;\s]+", normalize(value)):
        if part.isdigit():
            rows.add(int(part))
    return rows


def runtime_dir(config_path: Path) -> Path:
    return (config_path.parent.parent / "runtime").resolve()


def sender_profile(config_dir: Path, config: dict[str, Any]) -> dict[str, Any]:
    sender = config["sender"]
    profiles_path = resolve_path(config_dir, sender["profiles_path"])
    profiles = load_json(profiles_path)
    profile_name = sender["profile"]
    if profile_name not in profiles:
        raise RuntimeError(f"Unknown sender profile: {profile_name}")
    return profiles[profile_name]


def images_for_stage(config_dir: Path, config: dict[str, Any], stage: int) -> list[Path]:
    assets = config.get("assets", {})
    if stage == 1:
        values = assets.get("first_images", [])
    elif stage == 2:
        values = assets.get("second_images", assets.get("first_images", []))
    else:
        values = assets.get("third_images", [])
    paths = [resolve_path(config_dir, str(value)) for value in values]
    missing = [str(path) for path in paths if not path.exists()]
    if missing:
        raise RuntimeError("Missing product image(s): " + ", ".join(missing))
    return paths


def task_from_config(config: dict[str, Any]) -> dict[str, Any]:
    message = config.get("message", {})
    product_items = [normalize(item) for item in message.get("product_items", []) if normalize(item)]
    product_lines = []
    for item in product_items:
        words = [word for word in re.split(r"[^a-z0-9]+", item.lower()) if len(word) > 2]
        product_lines.append(
            {
                "id": compact(item) or item,
                "name": item,
                "keywords": words + [item.lower()],
                "angle": normalize(message.get("product_label")) or "vehicle accessory upgrade",
                "signals": [],
                "business_types": [],
            }
        )
    return {
        "vehicle_model": message.get("vehicle_model") or message.get("product_label") or "vehicle",
        "product_name": message.get("product_label", ""),
        "product_focus": message.get("product_label", ""),
        "product_family": message.get("product_label", ""),
        "campaign_angle": message.get("product_label", ""),
        "price_hint": message.get("price_hint", ""),
        "max_products": 3,
        "product_lines": product_lines,
    }


def render_html(config_dir: Path, config: dict[str, Any], lines: dict[str, str], image_count: int) -> str:
    shell = resolve_path(config_dir, config["templates"]["shell_path"]).read_text(encoding="utf-8")
    signature = resolve_path(config_dir, config["templates"]["signature_path"]).read_text(encoding="utf-8")
    variables = load_json(resolve_path(config_dir, config["templates"]["variables_path"]))
    image_style = variables.get("productImageStyle", "max-width:420px; width:100%; height:auto;")
    image_alt = config.get("message", {}).get("product_label", variables.get("productImageAlt", "Product image"))
    prefix = variables.get("productImageCidPrefix", "stage_img_")
    image_block = "\n".join(
        f'<p><img src="cid:{prefix}{index}" alt="{image_alt}" style="{image_style}"></p>'
        for index in range(1, image_count + 1)
    )
    feature = f"<p>{lines.get('feature', '')}</p>" if lines.get("feature") else ""
    replacements = {
        "{{ greeting }}": lines.get("greeting", "Hello,"),
        "{{ intro }}": lines.get("intro", ""),
        "{{ company_line_block }}": "",
        "{{ feature }}": feature,
        "{{ fit_line }}": lines.get("fit_line", ""),
        "{{ cta }}": lines.get("cta", ""),
        "{{ image_block }}": image_block,
        "{{ company_signature_block }}": signature,
        "{{ product_image_alt }}": image_alt,
    }
    rendered = shell
    for key, value in replacements.items():
        rendered = rendered.replace(key, value)
    return rendered


def build_email(config_dir: Path, config: dict[str, Any], sheet: Any, header_map: dict[str, int], row: int, stage: int) -> tuple[str, str, list[Path], dict[str, Any]]:
    task = task_from_config(config)
    customer = OutreachCustomer(
        company=cell_value(sheet, row, header_map, "company"),
        website=cell_value(sheet, row, header_map, "website"),
        email=cell_value(sheet, row, header_map, "email"),
        country=cell_value(sheet, row, header_map, "country"),
        row_index=row,
    )
    site_info = analyze_site(
        customer.website,
        company=customer.company,
        country=customer.country,
        task=task,
    )
    product_info = choose_products(site_info, task)
    variant_index = max(row - 2, 0)
    lines = build_lines(stage, customer, site_info, product_info, variant_index, task)
    subject = subject_for(stage, customer, product_info, site_info, variant_index, task)
    images = images_for_stage(config_dir, config, stage)
    html = render_html(config_dir, config, lines, len(images))
    return subject, html, images, {"site_info": site_info, "product_info": product_info}


def attach_inline_image(message: MIMEMultipart, cid: str, path: Path) -> None:
    with path.open("rb") as handle:
        image = MIMEImage(handle.read())
    image.add_header("Content-ID", f"<{cid}>")
    image.add_header("Content-Disposition", "inline", filename=path.name)
    message.attach(image)


def send_email(config_dir: Path, config: dict[str, Any], profile: dict[str, Any], to_email: str, subject: str, html: str, images: list[Path], stage: int) -> None:
    variables = load_json(resolve_path(config_dir, config["templates"]["variables_path"]))
    message = MIMEMultipart("related")
    message["Subject"] = subject
    message["From"] = profile["user"]
    message["To"] = to_email
    alternative = MIMEMultipart("alternative")
    alternative.attach(MIMEText(config.get("message", {}).get("plain_fallback", "Please view this email in HTML format."), "plain", "utf-8"))
    alternative.attach(MIMEText(html, "html", "utf-8"))
    message.attach(alternative)
    prefix = variables.get("productImageCidPrefix", "stage_img_")
    for index, image in enumerate(images, start=1):
        attach_inline_image(message, f"{prefix}{index}", image)
    logo_path = resolve_path(config_dir, config["assets"]["logo_path"])
    if logo_path.exists():
        attach_inline_image(message, variables.get("logoCid", "logo_img"), logo_path)
    with smtplib.SMTP_SSL(profile["host"], int(profile["port"]), timeout=30) as server:
        server.login(profile["user"], profile["password"])
        server.sendmail(profile["user"], [to_email], message.as_string())


def main() -> int:
    args = parse_args()
    config_path = Path(args.config).resolve()
    config_dir = config_path.parent
    config = load_json(config_path)
    rt_dir = runtime_dir(config_path)
    rt_dir.mkdir(parents=True, exist_ok=True)
    state_path = rt_dir / f"{config['campaign'].get('state_label', config['campaign'].get('name', 'campaign'))}_send_state.json"
    log_path = rt_dir / f"{config['campaign'].get('state_label', config['campaign'].get('name', 'campaign'))}_send_log.txt"
    state = load_json(state_path) if state_path.exists() else {"sent": [], "failed": []}
    workbook_path = Path(config["workbook"]["path"])
    workbook = load_workbook(workbook_path)
    sheet = workbook[config["campaign"].get("sheet_name", "Sheet1")] if config["campaign"].get("sheet_name") in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    header_map = map_headers(headers, field_aliases(config))
    required = ["email", "first_time"]
    missing = [key for key in required if key not in header_map]
    if missing:
        raise RuntimeError("Missing required column(s): " + ", ".join(missing))
    targets = parse_target_emails(args.target_emails)
    target_rows = parse_target_rows(args.target_rows)
    profile = sender_profile(config_dir, config)
    stage = 1
    processed = 0
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    today = datetime.now().strftime("%Y-%m-%d")
    for row in range(2, sheet.max_row + 1):
        email = cell_value(sheet, row, header_map, "email").lower()
        if not email:
            continue
        if targets and email not in targets:
            continue
        if target_rows and row not in target_rows:
            continue
        if not args.force_targets and cell_value(sheet, row, header_map, "first_time"):
            continue
        try:
            subject, html, images, analysis = build_email(config_dir, config, sheet, header_map, row, stage)
            if not images:
                raise RuntimeError("No product image configured")
            if args.test_to:
                send_email(config_dir, config, profile, args.test_to, subject, html, images, stage)
                to_address = args.test_to
            else:
                send_email(config_dir, config, profile, email, subject, html, images, stage)
                to_address = email
                sheet.cell(row, header_map["first_time"]).value = today
                workbook.save(workbook_path)
            item = {
                "row": row,
                "email": email,
                "to": to_address,
                "subject": subject,
                "time": now,
                "images": [str(path) for path in images],
                "selected_items": analysis["product_info"].get("items", []),
            }
            state.setdefault("sent", []).append(item)
            log_path.open("a", encoding="utf-8").write(f"[{now}] SENT {to_address} row={row} subject={subject} images={len(images)}\n")
            processed += 1
        except Exception as exc:
            state.setdefault("failed", []).append({"row": row, "email": email, "time": now, "error": str(exc)})
            save_json(state_path, state)
            log_path.open("a", encoding="utf-8").write(f"[{now}] FAILED {email} row={row} error={exc}\n")
            workbook.close()
            raise
        save_json(state_path, state)
        if args.limit and processed >= args.limit:
            break
        if not args.test_to and processed:
            delay_min = int(config["campaign"].get("delay_min_seconds", 0))
            delay_max = int(config["campaign"].get("delay_max_seconds", delay_min))
            if delay_max > 0:
                time.sleep(max(delay_min, delay_max))
    workbook.close()
    if processed == 0:
        raise RuntimeError("No matching target rows were sent")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
